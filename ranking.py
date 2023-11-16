import datetime
import math
import os
import time
import zipfile
from pathlib import Path
import numpy as np

import dateutil.relativedelta
import numpy as np
import pandas as pd
import win32com.client
from openpyxl import load_workbook  # Per caricare un libro
from openpyxl.styles import numbers  # Per cambiare i formati dei numeri
from openpyxl.styles import (Alignment, Font,  # Per cambiare lo stile
                             PatternFill)
from classes.metodi_ranking import Metodi_ranking


class Ranking():
    """
    BPPB è passata da metodo singolo a metodo doppio
    BPL è passata da metodo singolo a metodo doppio
    CRV usa il metodo lineare
    RIPA usa il metodo doppio
    RAI usa il metodo doppio
    """

    def __init__(self, intermediario):
        """
        Arguments:
            intermediario {str} - intermediario a cui è destinata l'analisi
        """
        self.intermediario = intermediario
        with open('docs/t1.txt') as f:
            t1 = f.read()
        t1 = datetime.datetime.strptime(t1, '%Y-%m-%d').strftime("%d/%m/%Y")
        self.t1 = t1
        # self.t0_3Y = (
        #     datetime.datetime.strptime(self.t1, '%d/%m/%Y') - dateutil.relativedelta.relativedelta(days=-1, years=+3)
        # ).strftime("%d/%m/%Y") # data iniziale tre anni fa
        self.t0_3Y = np.datetime64(
            datetime.datetime.strptime(self.t1, '%d/%m/%Y') - dateutil.relativedelta.relativedelta(days=-1, years=+3)
        ) # data iniziale tre anni fa

        # self.t0_1Y = (
        #     datetime.datetime.strptime(self.t1, '%d/%m/%Y') - dateutil.relativedelta.relativedelta(days=-1, years=+1)
        # ).strftime("%d/%m/%Y") # data iniziale un anno fa
        self.t0_1Y = np.datetime64(
            datetime.datetime.strptime(self.t1, '%d/%m/%Y') - dateutil.relativedelta.relativedelta(days=-1, years=+1)
        ) # data iniziale un anno fa

        # Directories
        directory = Path().cwd()
        self.directory = directory
        self.directory_output_liste = self.directory.joinpath('docs', 'export_liste_from_Q')
        self.directory_input_liste_best = self.directory.joinpath('docs', 'import_liste_best_into_Q')
        self.file_catalogo = 'catalogo_fondi.xlsx'
        self.file_completo = 'completo.csv'
        self.file_ranking_bw = 'ranking_bw.csv'
        self.file_ranking = 'ranking.xlsx'
        self.file_zip = 'rank.zip'

        match intermediario:
            # TODO: estendi il metodo classifica con linearizzazione anche a IR e PERF
            # più in generale gli indicatori che possono usare il metodo classifica con linearizzazione sono quelli
            # che possono usare la funzione sottostnte chiamata classifica. Stesso discorso per doppio con linearizzazione.
            # da utilizzare. Andrà a sostituire self.metodo. Questo permette di generalizzare ulteriormente questo file
            # e CRV avrà il metodo doppio alla loro maniera per le classi a benchmark inclusa la liqudità, e il metodo
            # linearizzazione per le altre.
            # TODO: togli self.metodo qui sotto, e sistema tutti i metodi che dipendono da esso
            case 'BPPB':
                self.metodo = 'doppio'
                self.metodi = {
                    'AZ_EUR': 'doppio', 'AZ_NA': 'doppio', 'AZ_PAC': 'doppio', 'AZ_EM': 'doppio',
                    'OBB_EUR_BT': 'doppio', 'OBB_EUR_MLT': 'doppio', 'OBB_EUR_CORP': 'doppio',
                    'OBB_GLOB': 'doppio', 'OBB_EM': 'doppio', 'OBB_HY': 'doppio', 'FLEX_BVOL': 'classifica',
                    'FLEX_MAVOL': 'classifica', 'OPP': 'classifica', 'LIQ': 'doppio',
                }
                self.soluzioni = {
                    'LIQ' : 1, 'OBB_EUR_BT' : 1, 'OBB_EUR_MLT' : 1, 'OBB_EUR_CORP' : 1, 'OBB_GLOB' : 1, 'OBB_EM' : 1,
                    'OBB_HY' : 1,
                    'AZ_EUR' : 3, 'AZ_NA' : 3, 'AZ_PAC' : 3, 'AZ_EM' : 3, 
                }
                self.classi_metodo_singolo = {
                    'AZ_EUR' : 'Az. Europa', 'AZ_NA' : 'Az. USA', 'AZ_PAC' : 'Az. Pacifico', 'AZ_EM' : 'Az. paesi emerg. Mondo',
                    'OBB_EUR_BT' : 'Obblig. Euro breve term.', 'OBB_EUR_MLT' : 'Obblig. Euro all maturities',
                    'OBB_EUR_CORP' : 'Obblig. Euro corporate', 'OBB_GLOB' : 'Obblig. globale', 'OBB_EM' : 'Obblig. Paesi Emerg.',
                    'OBB_HY' : 'Obblig. globale high yield'
                }
                self.classi_metodo_doppio = {
                    'LIQ' : 'Monetari Euro', 'OBB_EUR_BT' : 'Obblig. Euro breve term.', 'OBB_EUR_MLT' : 'Obblig. Euro all maturities', 
                    'OBB_EUR_CORP' : 'Obblig. Euro corporate', 'OBB_GLOB' : 'Obblig. globale', 'OBB_EM' : 'Obblig. Paesi Emerg.', 
                    'OBB_HY' : 'Obblig. globale high yield', 'AZ_EUR' : 'Az. Europa', 'AZ_NA' : 'Az. USA', 'AZ_PAC' : 'Az. Pacifico', 
                    'AZ_EM' : 'Az. paesi emerg. Mondo'
                }
                self.anni_detenzione = 3
                self.IR_TEV = [
                    'AZ_EUR', 'AZ_NA', 'AZ_PAC', 'AZ_EM', 'OBB_EUR_BT', 'OBB_EUR_MLT', 'OBB_EUR_CORP', 'OBB_GLOB',
                    'OBB_EM', 'OBB_HY'
                ]
                self.SOR_DSR = ['FLEX_BVOL', 'FLEX_MAVOL']
                self.SHA_VOL = ['OPP']
                self.PER_VOL = ['LIQ']
            case 'BPL':
                self.metodo = 'doppio'
                self.metodi = {
                    'AZ_EUR': 'doppio', 'AZ_NA': 'doppio', 'AZ_PAC': 'doppio', 'AZ_EM': 'doppio', 'AZ_GLOB': 'doppio',
                    'OBB_EUR_BT': 'doppio', 'OBB_EUR_MLT': 'doppio', 'OBB_EUR': 'doppio', 'OBB_EUR_CORP': 'doppio',
                    'OBB_GLOB': 'doppio', 'OBB_USA': 'doppio', 'OBB_EM': 'doppio', 'OBB_HY': 'doppio', 'BIL_MBVOL': 'classifica', 
                    'BIL_AVOL': 'classifica', 'FLEX_PR': 'classifica', 'FLEX_DIN': 'classifica', 'OPP': 'classifica', 'LIQ': 'doppio',
                    'LIQ_FOR': 'classifica',
                }
                self.soluzioni = {
                    'LIQ' : 3, 'OBB_EUR_BT' : 3, 'OBB_EUR_MLT' : 3, 'OBB_EUR' : 3, 'OBB_EUR_CORP' : 3, 'OBB_GLOB' : 3,
                    'OBB_USA' : 3, 'OBB_EM' : 3, 'OBB_HY' : 3, 'AZ_EUR' : 3, 'AZ_NA' : 3, 'AZ_PAC' : 3, 'AZ_EM' : 3,
                    'AZ_GLOB' : 3, 
                }
                self.classi_metodo_singolo = {
                    'AZ_EUR' : 'Az. Europa', 'AZ_NA' : 'Az. USA', 'AZ_PAC' : 'Az. Pacifico', 'AZ_EM' : 'Az. paesi emerg. Mondo', 
                    'AZ_GLOB' : 'Az. globale', 'OBB_EUR_BT' : 'Obblig. Euro breve term.',
                    'OBB_EUR_MLT' : 'Obblig. Euro all maturities', 'OBB_EUR' : 'Obblig. Europa', 'OBB_EUR_CORP' : 'Obblig. Euro corporate',
                    'OBB_GLOB' : 'Obblig. globale', 'OBB_USA' : 'Obblig. Dollaro US all mat', 'OBB_EM' : 'Obblig. Paesi Emerg.',
                    'OBB_HY' : 'Obblig. globale high yield'
                }
                self.classi_metodo_doppio = {
                    'LIQ' : 'Monetari Euro', 'OBB_EUR_BT' : 'Obblig. Euro breve term.', 'OBB_EUR_MLT' : 'Obblig. Euro all maturities', 
                    'OBB_EUR' : 'Obblig. Europa', 'OBB_EUR_CORP' : 'Obblig. Euro corporate', 'OBB_GLOB' : 'Obblig. globale', 
                    'OBB_USA' : 'Obblig. Dollaro US all mat', 'OBB_EM' : 'Obblig. Paesi Emerg.', 'OBB_HY' : 'Obblig. globale high yield', 
                    'AZ_EUR' : 'Az. Europa', 'AZ_NA' : 'Az. USA', 'AZ_PAC' : 'Az. Pacifico', 'AZ_EM' : 'Az. paesi emerg. Mondo', 
                    'AZ_GLOB' : 'Az. globale'
                }
                self.anni_detenzione = 5
                self.IR_TEV = [
                    'AZ_EUR', 'AZ_NA', 'AZ_PAC', 'AZ_EM', 'AZ_GLOB', 'OBB_EUR_BT', 'OBB_EUR_MLT', 'OBB_EUR', 'OBB_EUR_CORP',
                    'OBB_GLOB', 'OBB_USA', 'OBB_EM', 'OBB_HY'
                ]
                self.SOR_DSR = ['BIL_MBVOL', 'BIL_AVOL', 'FLEX_PR', 'FLEX_DIN']
                self.SHA_VOL = ['OPP']
                self.PER_VOL = ['LIQ', 'LIQ_FOR']
            case 'CRV':
                self.metodo = 'doppio'
                self.metodi = {'AZ_EUR': 'doppio_con_linearizzazione', 'AZ_NA': 'doppio_con_linearizzazione', 'AZ_PAC': 'doppio_con_linearizzazione', 
                    'AZ_EM': 'doppio_con_linearizzazione', 'AZ_GLOB': 'doppio_con_linearizzazione', 'OBB_EUR_BT': 'doppio_con_linearizzazione', 
                    'OBB_EUR_MLT': 'doppio_con_linearizzazione', 'OBB_EUR_CORP': 'doppio_con_linearizzazione', 'OBB_GLOB': 'doppio_con_linearizzazione', 
                    'OBB_EM': 'doppio_con_linearizzazione', 'OBB_HY': 'doppio_con_linearizzazione', 'FLEX_PR': 'classifica_con_linearizzazione', 
                    'FLEX_DIN': 'classifica_con_linearizzazione', 'OPP': 'classifica_con_linearizzazione', 'LIQ': 'doppio_con_linearizzazione', 
                }
                self.soluzioni = {
                    'LIQ' : 4, 'OBB_EUR_BT' : 4, 'OBB_EUR_MLT' : 4, 'OBB_EUR_CORP' : 4, 'OBB_GLOB' : 4, 'OBB_EM' : 4,
                    'OBB_HY' : 4, 'AZ_EUR' : 4, 'AZ_NA' : 4, 'AZ_PAC' : 4, 'AZ_EM' : 4, 'AZ_GLOB' : 4,
                }
                self.classi_metodo_singolo = None
                self.classi_metodo_doppio = {
                    'LIQ' : 'Monetari Euro', 'OBB_EUR_BT' : 'Obblig. Euro breve term.', 'OBB_EUR_MLT' : 'Obblig. Euro all maturities', 
                    'OBB_EUR_CORP' : 'Obblig. Euro corporate', 'OBB_GLOB' : 'Obblig. globale', 'OBB_EM' : 'Obblig. Paesi Emerg.',
                    'OBB_HY' : 'Obblig. globale high yield', 'AZ_EUR' : 'Az. Europa', 'AZ_NA' : 'Az. USA', 'AZ_PAC' : 'Az. Pacifico',
                    'AZ_EM' : 'Az. paesi emerg. Mondo', 'AZ_GLOB' : 'Az. globale', 
                }
                self.anni_detenzione = 3
                self.IR_TEV = [
                    'AZ_EUR', 'AZ_NA', 'AZ_PAC', 'AZ_EM', 'AZ_GLOB', 'OBB_EUR_BT', 'OBB_EUR_MLT', 'OBB_EUR_CORP', 'OBB_GLOB',
                    'OBB_EM', 'OBB_HY',
                ]
                self.SOR_DSR = ['FLEX_PR', 'FLEX_DIN']
                self.SHA_VOL = ['OPP']
                self.PER_VOL = ['LIQ']
            case 'RIPA':
                self.metodo = 'doppio'
                self.metodi = {
                    'AZ_EUR' : 'doppio', 'AZ_NA' : 'doppio', 'AZ_PAC' : 'doppio', 'AZ_EM' : 'doppio', 
                    'AZ_GLOB' : 'doppio', 'AZ_BIO' : 'doppio', 'AZ_BDC' : 'doppio', 'AZ_FIN' : 'doppio', 
                    'AZ_AMB' : 'doppio', 'AZ_IMM' : 'doppio', 'AZ_IND' : 'doppio', 'AZ_ECO' : 'doppio', 
                    'AZ_SAL' : 'doppio', 'AZ_SPU' : 'doppio', 'AZ_TEC' : 'doppio', 'AZ_TEL' : 'doppio', 
                    'AZ_ORO' : 'doppio', 'AZ_BEAR' : 'doppio', 
                    'OBB_EUR_BT' : 'doppio', 'OBB_EUR_MLT' : 'doppio', 'OBB_EUR' : 'doppio', 'OBB_EUR_CORP' : 'doppio', 
                    'OBB_GLOB' : 'doppio', 'OBB_USA' : 'doppio', 'OBB_JAP' : 'doppio', 'OBB_EM' : 'doppio', 'OBB_HY' : 'doppio', 
                    'FLEX_PR': 'classifica', 'FLEX_DIN': 'classifica', 
                    'COMM': 'classifica', 'PERF_ASS': 'classifica', 
                    'LIQ': 'doppio', 
                }
                self.soluzioni = {
                    'LIQ' : 4, 'OBB_EUR_BT' : 4, 'OBB_EUR_MLT' : 4, 'OBB_EUR' : 4, 'OBB_EUR_CORP' : 4, 'OBB_GLOB' : 4,
                    'OBB_USA' : 4, 'OBB_JAP' : 4, 'OBB_EM' : 4, 'OBB_HY' : 4, 'AZ_EUR' : 4, 'AZ_NA' : 4, 'AZ_PAC' : 4,
                    'AZ_EM' : 4, 'AZ_GLOB' : 4, 'AZ_BIO' : 4, 'AZ_BDC' : 4, 'AZ_FIN' : 4, 'AZ_AMB' : 4, 'AZ_IMM' : 4,
                    'AZ_IND' : 4, 'AZ_ECO' : 4, 'AZ_SAL' : 4, 'AZ_SPU' : 4, 'AZ_TEC' : 4, 'AZ_TEL' : 4, 'AZ_ORO' : 4,
                    'AZ_BEAR' : 4, 
                }
                self.classi_metodo_singolo = None
                self.classi_metodo_doppio = {
                    'LIQ' : 'Monetari Euro', 'OBB_EUR_BT' : 'Obblig. Euro breve term.', 'OBB_EUR_MLT' : 'Obblig. Euro all maturities', 
                    'OBB_EUR' : 'Obblig. Europa', 'OBB_EUR_CORP' : 'Obblig. Euro corporate', 'OBB_GLOB' : 'Obblig. globale', 
                    'OBB_USA' : 'Obblig. Dollaro US all mat', 'OBB_JAP' : 'Obblig. Yen', 'OBB_EM' : 'Obblig. Paesi Emerg.', 
                    'OBB_HY' : 'Obblig. globale high yield', 'AZ_EUR' : 'Az. Europa', 'AZ_NA' : 'Az. USA', 
                    'AZ_PAC' : 'Az. Pacifico', 'AZ_EM' : 'Az. paesi emerg. Mondo', 'AZ_GLOB' : 'Az. globale', 'AZ_BIO' : 'Az. Biotech', 
                    'AZ_BDC' : 'Az. beni di consumo', 'AZ_FIN' : 'Az. servizi finanziari', 'AZ_AMB' : 'Az. ambiente', 
                    'AZ_IMM' : 'Az. real estate Mondo', 'AZ_IND' : 'Az. industria', 'AZ_ECO' : 'Az. energia materie prime oro', 
                    'AZ_SAL' : 'Az. salute - farmaceutico', 'AZ_SPU' : 'Az. Servizi di pubblica utilita', 'AZ_TEC' : 'Az. tecnologia', 
                    'AZ_TEL' : 'Az. telecomunicazioni', 'AZ_ORO' : 'Az. Oro', 'AZ_BEAR' : 'Az. Bear',
                }
                self.anni_detenzione = 3
                self.IR_TEV = [
                    'OBB_EUR_BT', 'OBB_EUR_MLT', 'OBB_EUR', 'OBB_EUR_CORP', 'OBB_GLOB', 'OBB_USA', 'OBB_JAP', 'OBB_EM', 'OBB_HY', 
                    'AZ_EUR', 'AZ_NA', 'AZ_PAC', 'AZ_EM', 'AZ_GLOB', 'AZ_BIO', 'AZ_BDC', 'AZ_FIN', 'AZ_AMB', 'AZ_IMM', 'AZ_IND', 
                    'AZ_ECO', 'AZ_SAL', 'AZ_SPU', 'AZ_TEC', 'AZ_TEL', 'AZ_ORO', 'AZ_BEAR', 'FLEX_PR', 'FLEX_DIN', 
                ]
                # self.SOR_DSR = ['COMM', 'FLEX_PR', 'FLEX_DIN', 'PERF_ASS'] #SBAGLIATO
                self.SOR_DSR = ['COMM', 'PERF_ASS']
                self.SHA_VOL = []
                self.PER_VOL = ['LIQ']
            case 'RAI':
                self.metodo = 'doppio'
                self.metodi = {
                    'AZ_EUR': 'doppio', 'AZ_NA': 'doppio', 'AZ_PAC': 'doppio', 'AZ_EM': 'doppio', 'AZ_GLOB': 'doppio',
                    'OBB_EUR_BT': 'doppio', 'OBB_EUR_MLT': 'doppio', 'OBB_EUR': 'doppio', 'OBB_EUR_CORP': 'doppio',
                    'OBB_GLOB': 'doppio', 'OBB_USA': 'doppio', 'OBB_EM': 'doppio', 'OBB_HY': 'doppio', 
                    'BIL_PR': 'classifica', 'BIL_EQ': 'classifica', 'BIL_AGG': 'classifica', 'FLEX_PR': 'classifica', 
                    'FLEX_DIN': 'classifica', 'OPP': 'classifica', 'LIQ': 'doppio', 'LIQ_FOR': 'classifica',
                }
                self.soluzioni = {
                    'LIQ' : 4, 'OBB_EUR_BT' : 4, 'OBB_EUR_MLT' : 4, 'OBB_EUR_CORP' : 4, 'OBB_EUR' : 4, 'OBB_USA' : 4,
                    'OBB_GLOB' : 4, 'OBB_EM' : 4, 'OBB_HY' : 4, 'AZ_EUR' : 4, 'AZ_NA' : 4, 'AZ_PAC' : 4, 'AZ_EM' : 4,
                    'AZ_GLOB' : 4, 
                }
                self.classi_metodo_singolo = None
                self.classi_metodo_doppio = {
                    'LIQ' : 'Monetari Euro', 'OBB_EUR_BT' : 'Obblig. Euro breve term.', 'OBB_EUR_MLT' : 'Obblig. Euro all maturities', 
                    'OBB_EUR_CORP' : 'Obblig. Euro corporate', 'OBB_EUR' : 'Obblig. Europa', 'OBB_USA' : 'Obblig. Dollaro US all mat', 
                    'OBB_GLOB' : 'Obblig. globale', 'OBB_HY' : 'Obblig. globale high yield', 'OBB_EM' : 'Obblig. Paesi Emerg.', 
                    'AZ_EUR' : 'Az. Europa', 'AZ_NA' : 'Az. USA', 'AZ_PAC' : 'Az. Pacifico', 'AZ_EM' : 'Az. paesi emerg. Mondo', 
                    'AZ_GLOB' : 'Az. globale', 
                }
                self.anni_detenzione = None
                self.IR_TEV = [
                    'OBB_EUR_BT', 'OBB_EUR_MLT', 'OBB_EUR', 'OBB_USA', 'OBB_EUR_CORP', 'OBB_GLOB', 'OBB_EM', 'OBB_HY', 
                    'AZ_EUR', 'AZ_NA', 'AZ_PAC', 'AZ_EM', 'AZ_GLOB', 
                ]
                self.SOR_DSR = ['BIL_PR', 'BIL_EQ', 'BIL_AGG', 'FLEX_PR', 'FLEX_DIN']
                self.SHA_VOL = ['OPP']
                self.PER_VOL = ['LIQ', 'LIQ_FOR']
            case _:
                print('specifica un intermediario')
                quit()

    def attività(self):
        """
        Crea la colonna TEV ottenuta come rapporto tra alpha e IR, sia a 3 anni che ad 1 anno.
        Assegna ai fondi appartenenti alle classi direzionali, più la liquidità,
        un grado di attività tra semiattivo, attivo, molto attivo.
        L'etichetta verrà assegnata in base al superamento o meno di determinate soglie presenti nel dizionario soglie,
        condiviso tra tutti gli intermediari.
        """
        soglie = {
            'LIQ' : [0.0015, 0.01], 'OBB_EUR_BT' : [0.0075, 0.02], 'OBB_EUR_MLT' : [0.0125, 0.035], 'OBB_EUR' : [0.035, 0.065],
            'OBB_USA' : [0.035, 0.055], 'OBB_JAP' : [0.035, 0.06], 'OBB_EUR_CORP' : [0.01, 0.0275], 'OBB_GLOB' : [0.03, 0.06],
            'OBB_EM' : [0.045, 0.07], 'OBB_HY' : [0.04, 0.065], 'AZ_EUR' : [0.055, 0.10], 'AZ_NA' : [0.055, 0.10],
            'AZ_PAC' : [0.08, 0.12], 'AZ_EM' : [0.06, 0.14], 'AZ_GLOB' : [0.055, 0.10], 'AZ_BIO' : [0.08, 0.15],
            'AZ_BDC' : [0.08, 0.13], 'AZ_FIN' : [0.055, 0.12], 'AZ_AMB' : [0.08, 0.12], 'AZ_IMM' : [0.06, 0.10],
            'AZ_IND' : [0.055, 0.12], 'AZ_ECO' : [0.08, 0.14], 'AZ_SAL' : [0.055, 0.11], 'AZ_SPU' : [0.06, 0.12],
            'AZ_TEC' : [0.06, 0.14], 'AZ_TEL' : [0.05, 0.15], 'AZ_ORO' : [0.08, 0.15], 'AZ_BEAR' : [0.08, 0.15],
        }

        if self.metodo == 'singolo' or self.metodo == 'linearizzazione':
            return None
        df = pd.read_csv(self.file_completo, sep=";", decimal=',', index_col=None)
        df['data_di_avvio'] = pd.to_datetime(df['data_di_avvio'], dayfirst=True)
        if self.metodo == 'doppio':
            df.loc[(df['micro_categoria'].isin(list(self.classi_metodo_doppio.values()))) &
                (df['data_di_avvio'] < self.t0_3Y) & (df['Alpha 3 anni") fine mese'].notnull()), 'TEV_3Y'
            ] = df['Alpha 3 anni") fine mese'] / df['Info 3 anni") fine mese']
            for macro, micro in self.classi_metodo_doppio.items():
                df.loc[(df['micro_categoria']==micro) & (df['data_di_avvio'] < self.t0_3Y) &
                    (df['TEV_3Y'].notnull()), 'grado_gestione_3Y'
                ] = df.loc[(df['micro_categoria']==micro), 'TEV_3Y'].apply(
                    lambda x: 'semi_attivo' if x < soglie[macro][0] else 'attivo' if x < soglie[macro][1] else 'molto_attivo')
            df.loc[(df['micro_categoria'].isin(list(self.classi_metodo_doppio.values()))) &
                (df['data_di_avvio'] < self.t0_1Y) & (df['Alpha 1 anno fine mese'].notnull()), 'TEV_1Y'
            ] = df['Alpha 1 anno fine mese'] / df['Info 1 anno fine mese']
            for macro, micro in self.classi_metodo_doppio.items():
                df.loc[(df['micro_categoria']==micro) & (df['data_di_avvio'] < self.t0_1Y) &
                    (df['TEV_1Y'].notnull()), 'grado_gestione_1Y'
                ] = df.loc[(df['micro_categoria']==micro), 'TEV_1Y'].apply(
                    lambda x: 'semi_attivo' if x < soglie[macro][0] else 'attivo' if x < soglie[macro][1] else 'molto_attivo')
        df.to_csv(self.file_ranking_bw, sep=";", decimal=',', index=False)

    def indicatore_BS(self):
        """
        Metodo singolo
        1. Calcola l'indicatore B&S a 3 anni, correggendo l'IR per i costi spalmati sugli anni di detenzione medi di un fondo.

        Metodo doppio
        1. Calcola l'indicatore B&S a 3 anni, correggendo l'IR per i costi spalmati sugli anni di detenzione medi di un fondo.
        2. Calcola l'indicatore B&S a 1 anno, correggendo l'IR per i costi spalmati sugli anni di detenzione medi di un fondo.
        Formula v1 = IR - (IR * fee) / (anni_detenzione * alpha)
        Formula v2 = (IR * TEV - (fee / anni_detenzione)) / TEV
        Le colonne considerate ai fini del calcolo sono: 'Info 3 anni") fine mese', 'Alpha 3 anni") fine mese',
        'Info 1 anno fine mese', 'Alpha 1 anno fine mese', 'commissione'

        N.B. Raiffeisen non ha fornito un dato riassuntivo sugli anni di detenzione medi dei fondi.
             Nel loro caso sono costretto a portarmi dietro la colonna 'anni_detenzione' che indica gli anni di
             detenzione per singolo fondo.
        """
        if self.metodo == 'linearizzazione':
            return None
        df = pd.read_csv(self.file_ranking_bw, sep=";", decimal=',', index_col=None)
        df['data_di_avvio'] = pd.to_datetime(df['data_di_avvio'], dayfirst=True)
        if self.metodo == 'singolo':
            df.loc[
                (df['macro_categoria'].isin(list(self.classi_metodo_singolo.keys()))) & (df['data_di_avvio'] < self.t0_3Y), 'BS_3_anni'
            ] = df['Info 3 anni") fine mese'] - (df['Info 3 anni") fine mese'] * df['commissione']) / (int(self.anni_detenzione) * df['Alpha 3 anni") fine mese'])
        elif self.metodo == 'doppio':
            if self.intermediario != 'RAI':
                df.loc[
                    (df['macro_categoria'].isin(list(self.classi_metodo_doppio.keys()))) & (df['data_di_avvio'] < self.t0_3Y), 'BS_3_anni'
                ] = df['Info 3 anni") fine mese'] - (df['Info 3 anni") fine mese'] * df['commissione']) / (int(self.anni_detenzione) * df['Alpha 3 anni") fine mese'])
                df.loc[
                    (df['macro_categoria'].isin(list(self.classi_metodo_doppio.keys()))) & (df['data_di_avvio'] < self.t0_1Y), 'BS_1_anno'
                ] = df['Info 1 anno fine mese'] - (df['Info 1 anno fine mese'] * df['commissione']) / (int(self.anni_detenzione) * df['Alpha 1 anno fine mese'])
            elif self.intermediario == 'RAI': # Raiffeisen specifica gli anni di detenzione per singolo fondo
                if 'anni_detenzione' not in df.columns:
                    print('aggiungo la colonna "anni_detenzione"')
                    # Merge tra completo e catalogo per aggiungere la colonna anni_detenzione
                    df_catalogo = pd.read_excel(self.file_catalogo, index_col=None, usecols=['isin', 'anni_detenzione'])
                    df = pd.merge(left=df, right=df_catalogo, left_on='ISIN', right_on='isin')
                else:
                    print('la colonna "anni detenzione" esiste già')
                df.loc[
                    (df['macro_categoria'].isin(list(self.classi_metodo_doppio.keys()))) & (df['data_di_avvio'] < self.t0_3Y), 'BS_3_anni'
                ] = df['Info 3 anni") fine mese'] - (df['Info 3 anni") fine mese'] * df['commissione']) / (df['anni_detenzione'] * df['Alpha 3 anni") fine mese'])
                df.loc[
                    (df['macro_categoria'].isin(list(self.classi_metodo_doppio.keys()))) & (df['data_di_avvio'] < self.t0_1Y), 'BS_1_anno'
                ] = df['Info 1 anno fine mese'] - (df['Info 1 anno fine mese'] * df['commissione']) / (df['anni_detenzione'] * df['Alpha 1 anno fine mese'])
        df.to_csv(self.file_ranking_bw, sep=";", decimal=',', index=False)

    def calcolo_best_worst(self):
        """
        Metodo singolo
        1. Calcolo best e worst per le micro categorie contenute nelle macro categorie a benchmark, per fondi con più di tre anni
        e con indicatore B&S a tre anni presente, in base alla mediana.

        Metodo doppio
        1. Calcolo best e worst per le micro categorie contenute nelle macro categorie a benchmark, per fondi con più di tre anni
        e con indicatore B&S ad un anno presente, rispetto al grado di attività nel caso siano micro categorie direzionali, 
        in base alla mediana.
        2. Calcolo best e worst per le micro categorie contenute nelle macro categorie a benchmark, per fondi con più di un anno
        e con indicatore B&S ad un anno presente, rispetto al grado di attività nel caso siano micro categorie direzionali,
        in base al primo quartile.
        3. NON PIU' VALIDO I fondi con più di un anno di vita che sono best a 3 anni o best ad 1 anno, sono best, altrimenti worst.
        """
        
        if self.metodo == 'linearizzazione':
            return None

        df = pd.read_csv(self.file_ranking_bw, sep=";", decimal=',', index_col=None)
        df['data_di_avvio'] = pd.to_datetime(df['data_di_avvio'], dayfirst=True)

        if self.metodo == 'singolo':
            for macro in list(self.classi_metodo_singolo.keys()):
                for micro in df.loc[df['macro_categoria'] == macro, 'micro_categoria'].unique():
                    mediana = df.loc[
                        (df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_3Y)
                        & (df['BS_3_anni'].notnull()), 'BS_3_anni'].median()
                    df.loc[
                        (df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_3Y)
                        & (df['BS_3_anni'].notnull()), 'Best_Worst_3Y'
                    ] = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['fund_incept_dt'] < self.t0_3Y) &
                        (df['BS_3_anni'].notnull()), 'BS_3_anni'].apply(lambda x: 'worst' if x < mediana else 'best')
        elif self.metodo == 'doppio':
            for macro in list(self.classi_metodo_doppio.keys()):
                for micro in df.loc[df['macro_categoria'] == macro, 'micro_categoria'].unique():
                    # Chi sceglie la soluzione 4, vuole mescolare i tre diversi gradi di attività in ciascuna micro categoria
                    # Per Ripa si tratta della scelta definitiva, per altri intermediari si tratta di una condizione temporanea; 
                    # dopo il primo o i primi giri di ranking, queste realtà sceglieranno un ordinamento tra i tre disponibili.
                    if micro in list(self.classi_metodo_doppio.values()) and self.soluzioni[macro] != 4:
                        for grado in df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro), 'grado_gestione_3Y'].unique():
                            mediana = df.loc[
                                (df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_3Y)
                                & (df['BS_3_anni'].notnull()) & (df['grado_gestione_3Y'] == grado), 'BS_3_anni'].median()
                            df.loc[
                                (df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_3Y)
                                & (df['BS_3_anni'].notnull()) & (df['grado_gestione_3Y'] == grado), 'Best_Worst_3Y'
                            ] = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro)
                                & (df['data_di_avvio'] < self.t0_3Y) & (df['BS_3_anni'].notnull())
                                & (df['grado_gestione_3Y'] == grado), 'BS_3_anni'].apply(lambda x: 'worst' if x < mediana else 'best')
                        for grado in df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro), 'grado_gestione_1Y'].unique():
                            primo_quartile = df.loc[
                                (df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_1Y)
                                & (df['BS_1_anno'].notnull()) & (df['grado_gestione_1Y'] == grado), 'BS_1_anno'
                            ].quantile(q=0.75, interpolation='linear')
                            df.loc[
                                (df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_1Y)
                                & (df['BS_1_anno'].notnull()) & (df['grado_gestione_1Y'] == grado), 'Best_Worst_1Y'
                            ] = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro)
                                & (df['data_di_avvio'] < self.t0_1Y) & (df['BS_1_anno'].notnull())
                                & (df['grado_gestione_1Y'] == grado), 'BS_1_anno'].apply(lambda x: 'worst' if x < primo_quartile else 'best')
                    else:
                        mediana = df.loc[
                            (df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_3Y)
                            & (df['BS_3_anni'].notnull()), 'BS_3_anni'].median()
                        df.loc[
                            (df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_3Y)
                            & (df['BS_3_anni'].notnull()), 'Best_Worst_3Y'
                        ] = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_3Y)
                            & (df['BS_3_anni'].notnull()), 'BS_3_anni'].apply(lambda x: 'worst' if x < mediana else 'best')
                        primo_quartile = df.loc[
                            (df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_1Y)
                            & (df['BS_1_anno'].notnull()), 'BS_1_anno'].quantile(q=0.75, interpolation='linear')
                        df.loc[
                            (df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_1Y)
                            & (df['BS_1_anno'].notnull()), 'Best_Worst_1Y'
                        ] = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_1Y)
                            & (df['BS_1_anno'].notnull()), 'BS_1_anno'].apply(lambda x: 'worst' if x < primo_quartile else 'best')
            # df['Best_Worst'] = df['Best_Worst_3Y'].replace('worst', np.nan).fillna(df['Best_Worst_1Y'])
        df.to_csv(self.file_ranking_bw, sep=";", decimal=',', index=False)

    def ranking_per_grado(self):
        """
        Assegna un punteggio in ordine decrescente ai fondi delle micro categorie direzionali in base al loro indicatore corretto
        a 3 anni e ad 1 anno, discriminando in base al grado di gestione.

        Soluzione 1: semi-attivo; attivo; molto attivo;
        Soluzione 2: attivo; semi-attivo; molto attivo;
        Soluzione 3: attivo, semi-attivo; molto attivo;
        Soluzione 4: attivo, semi-attivo, molto attivo;
        """

        if self.metodo == 'singolo' or self.metodo == 'linearizzazione':
            return None
        elif self.metodo == 'doppio':
            df = pd.read_csv(self.file_ranking_bw, sep=";", decimal=',', index_col=None)
            df['data_di_avvio'] = pd.to_datetime(df['data_di_avvio'], dayfirst=True)

            for macro in list(self.classi_metodo_doppio.keys()):
                for micro in df.loc[df['macro_categoria'] == macro, 'micro_categoria'].unique():
                    if micro in list(self.classi_metodo_doppio.values()):
                        if self.soluzioni[macro] == 1:
                            # 3Y
                            for etichetta in ('best', 'worst'):
                                df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_3Y)
                                    & (df['BS_3_anni'].notnull()) & (df['grado_gestione_3Y'] == 'semi_attivo')
                                    & (df['Best_Worst_3Y'] == etichetta), 'ranking_per_grado_3Y'
                                ] = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_3Y)
                                    & (df['BS_3_anni'].notnull()) & (df['grado_gestione_3Y'] == 'semi_attivo')
                                    & (df['Best_Worst_3Y'] == etichetta), 'BS_3_anni'].rank(method='first', na_option='keep', ascending=False)
                                # Ottieni l'ultimo numero ordinale nella classificazione precedente
                                ultimo_elemento_ordinato_3Y = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & 
                                                                (df['data_di_avvio'] < self.t0_3Y) & (df['BS_3_anni'].notnull()) & 
                                                                (df['grado_gestione_3Y'] == 'semi_attivo') & 
                                                                (df['Best_Worst_3Y'] == etichetta), 'ranking_per_grado_3Y'].max()
                                # Quando non ci sono fondi semiattivi non c'è alcun ultimo numero ordinale.
                                if math.isnan(ultimo_elemento_ordinato_3Y): ultimo_elemento_ordinato_3Y = 0
                                df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_3Y)
                                    & (df['BS_3_anni'].notnull()) & (df['grado_gestione_3Y'] == 'attivo')
                                    & (df['Best_Worst_3Y'] == etichetta), 'ranking_per_grado_3Y'
                                ] = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_3Y)
                                    & (df['BS_3_anni'].notnull()) & (df['grado_gestione_3Y'] == 'attivo')
                                    & (df['Best_Worst_3Y'] == etichetta), 'BS_3_anni'].rank(method='first', na_option='keep', ascending=False) + ultimo_elemento_ordinato_3Y
                                # Ottieni l'ultimo numero ordinale nella classificazione precedente
                                ultimo_elemento_ordinato_3Y = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro)
                                                                & (df['data_di_avvio'] < self.t0_3Y) & (df['BS_3_anni'].notnull())
                                                                & (df['grado_gestione_3Y'] == 'attivo')
                                                                & (df['Best_Worst_3Y'] == etichetta), 'ranking_per_grado_3Y'].max()
                                # Quando non ci sono fondi semiattivi e attivi non c'è alcun ultimo numero ordinale.
                                if math.isnan(ultimo_elemento_ordinato_3Y): ultimo_elemento_ordinato_3Y = 0
                                df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_3Y)
                                    & (df['BS_3_anni'].notnull()) & (df['grado_gestione_3Y'] == 'molto_attivo')
                                    & (df['Best_Worst_3Y'] == etichetta), 'ranking_per_grado_3Y'
                                ] = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_3Y)
                                    & (df['BS_3_anni'].notnull()) & (df['grado_gestione_3Y'] == 'molto_attivo')
                                    & (df['Best_Worst_3Y'] == etichetta), 'BS_3_anni'].rank(method='first', na_option='keep', ascending=False) + ultimo_elemento_ordinato_3Y
                            # 1Y
                            for etichetta in ('best', 'worst'):
                                df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_1Y)
                                    & (df['BS_1_anno'].notnull()) & (df['grado_gestione_1Y'] == 'semi_attivo')
                                    & (df['Best_Worst_1Y'] == etichetta), 'ranking_per_grado_1Y'
                                ] = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_1Y)
                                    & (df['BS_1_anno'].notnull()) & (df['grado_gestione_1Y'] == 'semi_attivo')
                                    & (df['Best_Worst_1Y'] == etichetta), 'BS_1_anno'].rank(method='first', na_option='keep', ascending=False)
                                # Ottieni l'ultimo numero ordinale nella classificazione precedente
                                ultimo_elemento_ordinato_1Y = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & 
                                                                    (df['data_di_avvio'] < self.t0_1Y) & (df['BS_1_anno'].notnull()) & 
                                                                    (df['grado_gestione_1Y'] == 'semi_attivo') & 
                                                                    (df['Best_Worst_1Y'] == etichetta), 'ranking_per_grado_1Y'].max()
                                # Quando non ci sono fondi semiattivi non c'è alcun ultimo numero ordinale.
                                if math.isnan(ultimo_elemento_ordinato_1Y): ultimo_elemento_ordinato_1Y = 0
                                df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_1Y)
                                    & (df['BS_1_anno'].notnull()) & (df['grado_gestione_1Y'] == 'attivo')
                                    & (df['Best_Worst_1Y'] == etichetta), 'ranking_per_grado_1Y'
                                ] = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_1Y)
                                    & (df['BS_1_anno'].notnull()) & (df['grado_gestione_1Y'] == 'attivo')
                                    & (df['Best_Worst_1Y'] == etichetta), 'BS_1_anno'].rank(method='first', na_option='keep', ascending=False) + ultimo_elemento_ordinato_1Y
                                # Ottieni l'ultimo numero ordinale nella classificazione precedente
                                ultimo_elemento_ordinato_1Y = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & 
                                                                    (df['data_di_avvio'] < self.t0_1Y) & (df['BS_1_anno'].notnull()) & 
                                                                    (df['grado_gestione_1Y'] == 'attivo') & 
                                                                    (df['Best_Worst_1Y'] == etichetta), 'ranking_per_grado_1Y'].max()
                                # Quando non ci sono fondi attivi non c'è alcun ultimo numero ordinale.
                                if math.isnan(ultimo_elemento_ordinato_1Y): ultimo_elemento_ordinato_1Y = 0
                                df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_1Y)
                                    & (df['BS_1_anno'].notnull()) & (df['grado_gestione_1Y'] == 'molto_attivo')
                                    & (df['Best_Worst_1Y'] == etichetta), 'ranking_per_grado_1Y'
                                ] = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_1Y)
                                    & (df['BS_1_anno'].notnull()) & (df['grado_gestione_1Y'] == 'molto_attivo')
                                    & (df['Best_Worst_1Y'] == etichetta), 'BS_1_anno'].rank(method='first', na_option='keep', ascending=False) + ultimo_elemento_ordinato_1Y
                        elif self.soluzioni[macro] == 2:
                            # 3Y
                            for etichetta in ('best', 'worst'):
                                df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_3Y) & (df['BS_3_anni'].notnull()) & (df['grado_gestione_3Y'] == 'attivo') & (df['Best_Worst_3Y'] == etichetta), 'ranking_per_grado_3Y'] = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_3Y) & (df['BS_3_anni'].notnull()) & (df['grado_gestione_3Y'] == 'attivo') & (df['Best_Worst_3Y'] == etichetta), 'BS_3_anni'].rank(method='first', na_option='keep', ascending=False)
                                # Ottieni l'ultimo numero ordinale nella classificazione precedente
                                ultimo_elemento_ordinato_3Y = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & 
                                                                (df['data_di_avvio'] < self.t0_3Y) & (df['BS_3_anni'].notnull()) & 
                                                                (df['grado_gestione_3Y'] == 'attivo') & 
                                                                (df['Best_Worst_3Y'] == etichetta), 'ranking_per_grado_3Y'].max()
                                # Quando non ci sono fondi attivi non c'è alcun ultimo numero ordinale.
                                if math.isnan(ultimo_elemento_ordinato_3Y): ultimo_elemento_ordinato_3Y = 0
                                df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_3Y) & (df['BS_3_anni'].notnull()) & (df['grado_gestione_3Y'] == 'semi_attivo') & (df['Best_Worst_3Y'] == etichetta), 'ranking_per_grado_3Y'] = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_3Y) & (df['BS_3_anni'].notnull()) & (df['grado_gestione_3Y'] == 'semi_attivo') & (df['Best_Worst_3Y'] == etichetta), 'BS_3_anni'].rank(method='first', na_option='keep', ascending=False) + ultimo_elemento_ordinato_3Y
                                # Ottieni l'ultimo numero ordinale nella classificazione precedente
                                ultimo_elemento_ordinato_3Y = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & 
                                                                (df['data_di_avvio'] < self.t0_3Y) & (df['BS_3_anni'].notnull()) & 
                                                                (df['grado_gestione_3Y'] == 'semi_attivo') & 
                                                                (df['Best_Worst_3Y'] == etichetta), 'ranking_per_grado_3Y'].max()
                                # Quando non ci sono fondi semiattivi non c'è alcun ultimo numero ordinale.
                                if math.isnan(ultimo_elemento_ordinato_3Y): ultimo_elemento_ordinato_3Y = 0
                                df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_3Y) & (df['BS_3_anni'].notnull()) & (df['grado_gestione_3Y'] == 'molto_attivo') & (df['Best_Worst_3Y'] == etichetta), 'ranking_per_grado_3Y'] = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_3Y) & (df['BS_3_anni'].notnull()) & (df['grado_gestione_3Y'] == 'molto_attivo') & (df['Best_Worst_3Y'] == etichetta), 'BS_3_anni'].rank(method='first', na_option='keep', ascending=False) + ultimo_elemento_ordinato_3Y
                            # 1Y
                            for etichetta in ('best', 'worst'):
                                df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_1Y) & (df['BS_1_anno'].notnull()) & (df['grado_gestione_1Y'] == 'attivo') & (df['Best_Worst_1Y'] == etichetta), 'ranking_per_grado_1Y'] = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_1Y) & (df['BS_1_anno'].notnull()) & (df['grado_gestione_1Y'] == 'attivo') & (df['Best_Worst_1Y'] == etichetta), 'BS_1_anno'].rank(method='first', na_option='keep', ascending=False)
                                # Ottieni l'ultimo numero ordinale nella classificazione precedente
                                ultimo_elemento_ordinato_1Y = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & 
                                                                    (df['data_di_avvio'] < self.t0_1Y) & (df['BS_1_anno'].notnull()) & 
                                                                    (df['grado_gestione_1Y'] == 'attivo') & 
                                                                    (df['Best_Worst_1Y'] == etichetta), 'ranking_per_grado_1Y'].max()
                                # Quando non ci sono fondi attivi non c'è alcun ultimo numero ordinale.
                                if math.isnan(ultimo_elemento_ordinato_1Y): ultimo_elemento_ordinato_1Y = 0
                                df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_1Y) & (df['BS_1_anno'].notnull()) & (df['grado_gestione_1Y'] == 'semi_attivo') & (df['Best_Worst_1Y'] == etichetta), 'ranking_per_grado_1Y'] = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_1Y) & (df['BS_1_anno'].notnull()) & (df['grado_gestione_1Y'] == 'semi_attivo') & (df['Best_Worst_1Y'] == etichetta), 'BS_1_anno'].rank(method='first', na_option='keep', ascending=False) + ultimo_elemento_ordinato_1Y
                                # Ottieni l'ultimo numero ordinale nella classificazione precedente
                                ultimo_elemento_ordinato_1Y = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & 
                                                                    (df['data_di_avvio'] < self.t0_1Y) & (df['BS_1_anno'].notnull()) & 
                                                                    (df['grado_gestione_1Y'] == 'semi_attivo') & 
                                                                    (df['Best_Worst_1Y'] == etichetta), 'ranking_per_grado_1Y'].max()
                                # Quando non ci sono fondi semiattivi non c'è alcun ultimo numero ordinale.
                                if math.isnan(ultimo_elemento_ordinato_1Y): ultimo_elemento_ordinato_1Y = 0
                                df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_1Y) & (df['BS_1_anno'].notnull()) & (df['grado_gestione_1Y'] == 'molto_attivo') & (df['Best_Worst_1Y'] == etichetta), 'ranking_per_grado_1Y'] = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_1Y) & (df['BS_1_anno'].notnull()) & (df['grado_gestione_1Y'] == 'molto_attivo') & (df['Best_Worst_1Y'] == etichetta), 'BS_1_anno'].rank(method='first', na_option='keep', ascending=False) + ultimo_elemento_ordinato_1Y
                        elif self.soluzioni[macro] == 3:
                            # 3Y
                            for etichetta in ('best', 'worst'):
                                df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_3Y) & (df['BS_3_anni'].notnull()) & (df['grado_gestione_3Y'].isin(['attivo', 'semi_attivo'])) & (df['Best_Worst_3Y'] == etichetta), 'ranking_per_grado_3Y'] = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_3Y) & (df['BS_3_anni'].notnull()) & (df['grado_gestione_3Y'].isin(['attivo', 'semi_attivo'])) & (df['Best_Worst_3Y'] == etichetta), 'BS_3_anni'].rank(method='first', na_option='keep', ascending=False)
                                # Ottieni l'ultimo numero ordinale nella classificazione precedente
                                ultimo_elemento_ordinato_3Y = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & 
                                                                (df['data_di_avvio'] < self.t0_3Y) & (df['BS_3_anni'].notnull()) & 
                                                                (df['grado_gestione_3Y'].isin(['attivo', 'semi_attivo'])) & 
                                                                (df['Best_Worst_3Y'] == etichetta), 'ranking_per_grado_3Y'].max()
                                # Quando non ci sono fondi attivi o semiattivi non c'è alcun ultimo numero ordinale.
                                if math.isnan(ultimo_elemento_ordinato_3Y): ultimo_elemento_ordinato_3Y = 0
                                df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_3Y) & (df['BS_3_anni'].notnull()) & (df['grado_gestione_3Y'] == 'molto_attivo') & (df['Best_Worst_3Y'] == etichetta), 'ranking_per_grado_3Y'] = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_3Y) & (df['BS_3_anni'].notnull()) & (df['grado_gestione_3Y'] == 'molto_attivo') & (df['Best_Worst_3Y'] == etichetta), 'BS_3_anni'].rank(method='first', na_option='keep', ascending=False) + ultimo_elemento_ordinato_3Y
                            # 1Y
                            for etichetta in ('best', 'worst'):
                                df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_1Y) & (df['BS_1_anno'].notnull()) & (df['grado_gestione_1Y'].isin(['attivo', 'semi_attivo'])) & (df['Best_Worst_1Y'] == etichetta), 'ranking_per_grado_1Y'] = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_1Y) & (df['BS_1_anno'].notnull()) & (df['grado_gestione_1Y'].isin(['attivo', 'semi_attivo'])) & (df['Best_Worst_1Y'] == etichetta), 'BS_1_anno'].rank(method='first', na_option='keep', ascending=False)
                                # Ottieni l'ultimo numero ordinale nella classificazione precedente
                                ultimo_elemento_ordinato_1Y = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & 
                                                                    (df['data_di_avvio'] < self.t0_1Y) & (df['BS_1_anno'].notnull()) & 
                                                                    (df['grado_gestione_1Y'].isin(['attivo', 'semi_attivo'])) & 
                                                                    (df['Best_Worst_1Y'] == etichetta), 'ranking_per_grado_1Y'].max()
                                # Quando non ci sono fondi attivi o semiattivi non c'è alcun ultimo numero ordinale.
                                if math.isnan(ultimo_elemento_ordinato_1Y): ultimo_elemento_ordinato_1Y = 0
                                df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_1Y) & (df['BS_1_anno'].notnull()) & (df['grado_gestione_1Y'] == 'molto_attivo') & (df['Best_Worst_1Y'] == etichetta), 'ranking_per_grado_1Y'] = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_1Y) & (df['BS_1_anno'].notnull()) & (df['grado_gestione_1Y'] == 'molto_attivo') & (df['Best_Worst_1Y'] == etichetta), 'BS_1_anno'].rank(method='first', na_option='keep', ascending=False) + ultimo_elemento_ordinato_1Y
                        elif self.soluzioni[macro] == 4:
                            # 3Y
                            for etichetta in ('best', 'worst'):
                                df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_3Y) & (df['BS_3_anni'].notnull()) & (df['grado_gestione_3Y'].isin(['molto_attivo', 'attivo', 'semi_attivo'])) & (df['Best_Worst_3Y'] == etichetta), 'ranking_per_grado_3Y'] = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_3Y) & (df['BS_3_anni'].notnull()) & (df['grado_gestione_3Y'].isin(['molto_attivo', 'attivo', 'semi_attivo'])) & (df['Best_Worst_3Y'] == etichetta), 'BS_3_anni'].rank(method='first', na_option='keep', ascending=False)
                            # 1Y
                            for etichetta in ('best', 'worst'):
                                df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_1Y) & (df['BS_1_anno'].notnull()) & (df['grado_gestione_1Y'].isin(['molto_attivo', 'attivo', 'semi_attivo'])) & (df['Best_Worst_1Y'] == etichetta), 'ranking_per_grado_1Y'] = df.loc[(df['macro_categoria'] == macro) & (df['micro_categoria'] == micro) & (df['data_di_avvio'] < self.t0_1Y) & (df['BS_1_anno'].notnull()) & (df['grado_gestione_1Y'].isin(['molto_attivo', 'attivo', 'semi_attivo'])) & (df['Best_Worst_1Y'] == etichetta), 'BS_1_anno'].rank(method='first', na_option='keep', ascending=False)

            df.to_csv(self.file_ranking_bw, sep=";", decimal=',', index=False)

    def merge_completo_liste(self):
        """Aggiunge gli indici scaricati da Quantalys.it nel file completo
        Passando dal percorso Fondi -> Confronto, alcuni fondi estinti o assorbiti vengono eslusi dalla lista caricata.
        Questo porta ad avere i dati di un numero di fondi inferiore a quelli caricati nelle liste.
        # TODO: Devo verificare se i fondi che non sono stati scaricati da Quantalys, non siano stati classificati come best nel processo
        precedente altrimenti sorgerebbe un problema.
        """
        if self.metodo == 'linearizzazione'  or self.metodo == 'classifica':
            df = pd.read_csv(self.file_completo, sep=";", decimal=',', index_col=None)
            colonne = [
                'ISIN', 'valuta', 'nome', 'micro_categoria', 'macro_categoria', 'data_di_avvio', 'commissione',
                'SFDR',
            ]
        elif self.metodo == 'doppio' or self.metodo == 'doppio_con_linearizzazione':
            df = pd.read_csv(self.file_ranking_bw, sep=";", decimal=',', index_col=None)
            colonne = [
                'ISIN', 'valuta', 'nome', 'micro_categoria', 'macro_categoria', 'data_di_avvio', 'commissione',
                'SFDR', 'grado_gestione_3Y', 'grado_gestione_1Y', 'Best_Worst_3Y', 'Best_Worst_1Y', 'ranking_per_grado_3Y',
                'ranking_per_grado_1Y', 
            ]
        df = df[colonne]
        print('sto aggiungendo gli indici delle liste al file completo...\n')

        df['Information_Ratio_3Y'] = np.nan
        df['TEV_3Y'] = np.nan
        df['Information_Ratio_1Y'] = np.nan
        df['TEV_1Y'] = np.nan
        df['Sortino_3Y'] = np.nan
        df['DSR_3Y'] = np.nan
        df['Sortino_1Y'] = np.nan
        df['DSR_1Y'] = np.nan
        df['Perf_3Y'] = np.nan
        df['Vol_3Y'] = np.nan
        df['Perf_1Y'] = np.nan
        df['Vol_1Y'] = np.nan
        df['Sharpe_3Y'] = np.nan
        df['Sharpe_1Y'] = np.nan

        for filename in os.listdir(self.directory_output_liste):
            if filename[:-9] in self.IR_TEV:
                i = 10
                while True:
                    try:
                        lista_indici = pd.read_csv(self.directory_output_liste.joinpath(filename), sep = ';', header=0, skiprows=2, skipfooter=i, engine='python', encoding='unicode_escape')
                    except pd.errors.ParserError:
                        # se i fondi sono pochi la tabella delle correlazioni viene riempita al completo e le righe da saltare in fondo sono più di 14
                        i = i + 1
                    else:
                        if lista_indici.Nome.isna().any():
                            i = i + 1
                            continue
                        else:
                            break
                lista_indici = lista_indici.replace(',', '.', regex=True).astype({'Information ratio': float, 'TEV' : float})
                if filename[-6:-4] == '3Y':
                    df = df.merge(lista_indici, how='outer', left_on='ISIN', right_on='Codice ISIN')
                    df['Information_Ratio_3Y'] = df['Information ratio'].fillna(df['Information_Ratio_3Y'])
                    df['TEV_3Y'] = df['TEV'].fillna(df['TEV_3Y'])
                    df.drop(['Codice ISIN', 'Nome', 'Valuta', 'Information ratio', 'TEV'], axis=1, inplace=True)
                    df = df.astype({'Information_Ratio_3Y' : float, 'TEV_3Y' : float})
                elif filename[-6:-4] == '1Y':
                    df = df.merge(lista_indici, how='outer', left_on='ISIN', right_on='Codice ISIN')
                    df['Information_Ratio_1Y'] = df['Information ratio'].fillna(df['Information_Ratio_1Y'])
                    df['TEV_1Y'] = df['TEV'].fillna(df['TEV_1Y'])
                    df.drop(['Codice ISIN', 'Nome', 'Valuta', 'Information ratio', 'TEV'], axis=1, inplace=True)
                    df = df.astype({'Information_Ratio_1Y' : float, 'TEV_1Y' : float})
            elif filename[:-9] in self.SOR_DSR:
                i = 10
                while True:
                    try:
                        lista_indici = pd.read_csv(self.directory_output_liste.joinpath(filename), sep = ';', header=0, skiprows=2, skipfooter=i, engine='python', encoding='unicode_escape')
                    except pd.errors.ParserError:
                        # se i fondi sono pochi la tabella delle correlazioni viene riempita al completo e le righe da saltare in fondo sono più di 14
                        i = i + 1
                    else:
                        if lista_indici.Nome.isna().any():
                            i = i + 1
                            continue
                        else:
                            break
                lista_indici = lista_indici.replace(',', '.', regex=True).astype({'Sortino ratio': float, 'DSR' : float})
                if filename[-6:-4] == '3Y':
                    df = df.merge(lista_indici, how='outer', left_on='ISIN', right_on='Codice ISIN')
                    df['Sortino_3Y'] = df['Sortino ratio'].fillna(df['Sortino_3Y'])
                    df['DSR_3Y'] = df['DSR'].fillna(df['DSR_3Y'])
                    df.drop(['Codice ISIN', 'Nome', 'Valuta', 'Sortino ratio', 'DSR'], axis=1, inplace=True)
                elif filename[-6:-4] == '1Y':
                    df = df.merge(lista_indici, how='outer', left_on='ISIN', right_on='Codice ISIN')
                    df['Sortino_1Y'] = df['Sortino ratio'].fillna(df['Sortino_1Y'])
                    df['DSR_1Y'] = df['DSR'].fillna(df['DSR_1Y'])
                    df.drop(['Codice ISIN', 'Nome', 'Valuta', 'Sortino ratio', 'DSR'], axis=1, inplace=True)
            elif filename[:-9] in self.SHA_VOL:
                i = 10
                while True:
                    try:
                        lista_indici = pd.read_csv(self.directory_output_liste.joinpath(filename), sep = ';', header=0, skiprows=2, skipfooter=i, engine='python', encoding='unicode_escape')
                    except pd.errors.ParserError:
                        # se i fondi sono pochi la tabella delle correlazioni viene riempita al completo e le righe da saltare in fondo sono più di 14
                        i = i + 1
                    else:
                        if lista_indici.Nome.isna().any():
                            i = i + 1
                            continue
                        else:
                            break
                lista_indici = lista_indici.replace(',', '.', regex=True).astype({'Sharpe ratio': float, 'Volatilità' : float})
                #print(filename)
                if filename[-6:-4] == '3Y':
                    df = df.merge(lista_indici, how='outer', left_on='ISIN', right_on='Codice ISIN')
                    df['Sharpe_3Y'] = df['Sharpe ratio'].fillna(df['Sharpe_3Y'])
                    df['Vol_3Y'] = df['Volatilità'].fillna(df['Vol_3Y'])
                    df.drop(['Codice ISIN', 'Nome', 'Valuta', 'Sharpe ratio', 'Volatilità'], axis=1, inplace=True)
                elif filename[-6:-4] == '1Y':
                    df = df.merge(lista_indici, how='outer', left_on='ISIN', right_on='Codice ISIN')
                    df['Sharpe_1Y'] = df['Sharpe ratio'].fillna(df['Sharpe_1Y'])
                    df['Vol_1Y'] = df['Volatilità'].fillna(df['Vol_1Y'])
                    df.drop(['Codice ISIN', 'Nome', 'Valuta', 'Sharpe ratio', 'Volatilità'], axis=1, inplace=True)
            elif filename[:-9] in self.PER_VOL:
                i = 10
                while True:
                    try:
                        lista_indici = pd.read_csv(self.directory_output_liste.joinpath(filename), sep = ';', header=0, skiprows=2, skipfooter=i, engine='python', encoding='unicode_escape')
                    except pd.errors.ParserError:
                        # se i fondi sono pochi la tabella delle correlazioni viene riempita al completo e le righe da saltare in fondo sono più di 14
                        i = i + 1
                    else:
                        if lista_indici.Nome.isna().any():
                            i = i + 1
                            continue
                        else:
                            break
                lista_indici = lista_indici.replace(',', '.', regex=True).astype({'Perf Ann.': float, 'Volatilità' : float})
                #print(filename)
                if filename[-6:-4] == '3Y':
                    df = df.merge(lista_indici, how='outer', left_on='ISIN', right_on='Codice ISIN')
                    df['Perf_3Y'] = df['Perf Ann.'].fillna(df['Perf_3Y'])
                    df['Vol_3Y'] = df['Volatilità'].fillna(df['Vol_3Y'])
                    df.drop(['Codice ISIN', 'Nome', 'Valuta', 'Perf Ann.', 'Volatilità'], axis=1, inplace=True)
                elif filename[-6:-4] == '1Y':
                    df = df.merge(lista_indici, how='outer', left_on='ISIN', right_on='Codice ISIN')
                    df['Perf_1Y'] = df['Perf Ann.'].fillna(df['Perf_1Y'])
                    df['Vol_1Y'] = df['Volatilità'].fillna(df['Vol_1Y'])
                    df.drop(['Codice ISIN', 'Nome', 'Valuta', 'Perf Ann.', 'Volatilità'], axis=1, inplace=True)
        df.to_excel(self.file_ranking, index=False)

    def rank(self):
        # TODO : fallo all'interno di un wrapper come il metodo aggiunta_colonne
        """Crea il file di ranking con tanti fogli quante sono le macro asset class."""
        # Creazione file ranking diviso per macro
        print("sto facendo l'ordinamento dei fondi\n")
        df = pd.read_excel(self.file_ranking, index_col=None)
        df['data_di_avvio'] = pd.to_datetime(df['data_di_avvio'], dayfirst=True)
        if self.intermediario == 'RAI': # Raiffeisen specifica gli anni di detenzione per singolo fondo
            df_catalogo = pd.read_excel(self.file_catalogo, index_col=None, usecols=['isin', 'anni_detenzione'])
            # Merge tra completo e catalogo per aggiungere la colonna anni_detenzione
            df = pd.merge(left=df, right=df_catalogo, left_on='ISIN', right_on='isin')
        writer = pd.ExcelWriter(self.file_ranking,  engine='xlsxwriter') # pylint: disable=abstract-class-instantiated

        for macro in df.loc[:, 'macro_categoria'].unique():
            # Crea un foglio per ogni macro categoria
            foglio = df.loc[df['macro_categoria'] == macro].copy()
            metodo = self.metodi[macro]
            print(macro)
            print(metodo)
            if macro in self.IR_TEV:
                # Metodo best-worst singolo
                if self.metodo == 'singolo':
                    # Rank IR_1Y
                    foglio['ranking_IR_1Y'] = foglio.loc[
                        (foglio['data_di_avvio'] < self.t0_1Y) & (foglio['Information_Ratio_1Y'].notnull()), 'Information_Ratio_1Y'
                    ].rank(method='first', na_option='bottom', ascending=False)
                    # Quartile IR_1Y
                    foglio['quartile_IR_1Y'] = foglio.loc[
                        (foglio['data_di_avvio'] < self.t0_1Y) & (foglio['Information_Ratio_1Y'].notnull()), 'Information_Ratio_1Y'
                    ].apply(lambda x: 'best' if x > foglio['Information_Ratio_1Y'].quantile(0.25, interpolation = 'linear') else 'worst')
                    # Terzile IR_1Y
                    foglio['terzile_IR_1Y'] = foglio.loc[
                        (foglio['data_di_avvio'] < self.t0_1Y) & (foglio['Information_Ratio_1Y'].notnull()), 'Information_Ratio_1Y'
                    ].apply(lambda x: 'best' if x > foglio['Information_Ratio_1Y'].quantile(0.33, interpolation = 'linear') else 'worst')
                    # Creazione IR_corretto_1Y
                    foglio['IR_corretto_1Y'] = (
                        (df['Information_Ratio_1Y'] * (df['TEV_1Y'] / 100) ) - (df['commissione'] / self.anni_detenzione)) / (df['TEV_1Y'] / 100)
                    # Rank IR_corretto_1Y
                    foglio['ranking_IR_1Y_corretto'] = foglio.loc[
                        (foglio['data_di_avvio'] < self.t0_1Y) & (foglio['IR_corretto_1Y'].notnull()), 'IR_corretto_1Y'
                    ].rank(method='first', na_option='bottom', ascending=False)
                    # Quartile IR_1Y corretto
                    foglio['quartile_IR_corretto_1Y'] = foglio.loc[
                        (foglio['data_di_avvio'] < self.t0_1Y) & (foglio['IR_corretto_1Y'].notnull()), 'IR_corretto_1Y'
                    ].apply(lambda x: 'best' if x > foglio['IR_corretto_1Y'].quantile(0.25, interpolation = 'linear') else 'worst')
                    # Terzile IR_1Y corretto
                    foglio['terzile_IR_corretto_1Y'] = foglio.loc[
                        (foglio['data_di_avvio'] < self.t0_1Y) & (foglio['IR_corretto_1Y'].notnull()), 'IR_corretto_1Y'
                    ].apply(lambda x: 'best' if x > foglio['IR_corretto_1Y'].quantile(0.33, interpolation = 'linear') else 'worst')

                    # Rank IR_3Y
                    foglio['ranking_IR_3Y'] = foglio.loc[
                        (foglio['data_di_avvio'] < self.t0_3Y) & (foglio['Best_Worst'].notnull())
                        & (foglio['Information_Ratio_3Y'].notnull()), 'Information_Ratio_3Y'
                    ].rank(method='first', na_option='keep', ascending=False)
                    # Note
                    foglio.loc[(foglio['data_di_avvio'] < self.t0_3Y) & foglio['Best_Worst'].isnull(), 'note'] = 'Ha 3 anni, ma non è in classifica.'
                    foglio.loc[(foglio['data_di_avvio'] > self.t0_3Y) & foglio['Information_Ratio_3Y'].notnull(), 'note'] = 'Non ha 3 anni, ma possiede dati a tre anni.'
                    # Quartile IR_3Y
                    foglio['quartile_IR_3Y'] = foglio.loc[
                        (foglio['data_di_avvio'] < self.t0_3Y) & (foglio['Best_Worst'].notnull())
                        & (foglio['Information_Ratio_3Y'].notnull()), 'Information_Ratio_3Y'
                    ].apply(lambda x: 'best' if x > foglio['Information_Ratio_3Y'].quantile(0.25, interpolation = 'linear') else 'worst')
                    # Terzile IR_3Y
                    foglio['terzile_IR_3Y'] = foglio.loc[
                        (foglio['data_di_avvio'] < self.t0_3Y) & (foglio['Best_Worst'].notnull())
                        & (foglio['Information_Ratio_3Y'].notnull()), 'Information_Ratio_3Y'
                    ].apply(lambda x: 'best' if x > foglio['Information_Ratio_3Y'].quantile(0.33, interpolation = 'linear') else 'worst')
                    # Creazione IR_corretto_3Y
                    foglio['IR_corretto_3Y'] = (
                        (df['Information_Ratio_3Y'] * (df['TEV_3Y'] / 100)) - (df['commissione'] / self.anni_detenzione)) / (df['TEV_3Y'] / 100)
                    # Rank IR_corretto_3Y
                    foglio['ranking_IR_3Y_corretto'] = foglio.loc[
                        (foglio['data_di_avvio'] < self.t0_3Y) & (foglio['Best_Worst'].notnull())
                        & (foglio['IR_corretto_3Y'].notnull()), 'IR_corretto_3Y'
                    ].rank(method='first', na_option='bottom', ascending=False)
                    # Quartile IR_3Y corretto
                    foglio['quartile_IR_corretto_3Y'] = foglio.loc[
                        (foglio['data_di_avvio'] < self.t0_3Y) & (foglio['Best_Worst'].notnull())
                        & (foglio['IR_corretto_3Y'].notnull()), 'IR_corretto_3Y'
                    ].apply(lambda x: 'best' if x > foglio['IR_corretto_3Y'].quantile(0.25, interpolation = 'linear') else 'worst')
                    # Terzile IR_3Y corretto
                    foglio['terzile_IR_corretto_3Y'] = foglio.loc[
                        (foglio['data_di_avvio'] < self.t0_3Y) & (foglio['Best_Worst'].notnull())
                        & (foglio['IR_corretto_3Y'].notnull()), 'IR_corretto_3Y'
                    ].apply(lambda x: 'best' if x > foglio['IR_corretto_3Y'].quantile(0.33, interpolation = 'linear') else 'worst')
                    
                    # Ranking finale
                    foglio.loc[
                        (foglio['data_di_avvio'] < self.t0_3Y) & (foglio['Best_Worst'] == 'best') & (foglio['IR_corretto_3Y'].notnull())
                        & (foglio['micro_categoria'] == self.classi_metodo_singolo[macro]), 'ranking_finale'
                    ] = foglio.loc[
                        (foglio['data_di_avvio'] < self.t0_3Y) & (foglio['Best_Worst'] == 'best') & (foglio['IR_corretto_3Y'].notnull())
                        & (foglio['micro_categoria'] == self.classi_metodo_singolo[macro]), 'IR_corretto_3Y'
                        ].rank(method='first', na_option='bottom', ascending=False)
                    foglio.loc[
                        (foglio['data_di_avvio'] < self.t0_3Y) & (foglio['Best_Worst'] == 'best')
                        & (foglio['micro_categoria'] != self.classi_metodo_singolo[macro]), 'ranking_finale'
                    ] = foglio.loc[
                        (foglio['data_di_avvio'] < self.t0_3Y) & (foglio['Best_Worst'] == 'best')
                        & (foglio['micro_categoria'] != self.classi_metodo_singolo[macro]), 'IR_corretto_3Y'
                        ].rank(method='first', na_option='bottom', ascending=False) + foglio['ranking_finale'].max()
                    foglio.loc[
                        (foglio['data_di_avvio'] < self.t0_3Y) & (foglio['Best_Worst'] == 'worst')
                        & (foglio['micro_categoria'] == self.classi_metodo_singolo[macro]), 'ranking_finale'
                    ] = foglio.loc[
                        (foglio['data_di_avvio'] < self.t0_3Y) & (foglio['Best_Worst'] == 'worst')
                        & (foglio['micro_categoria'] == self.classi_metodo_singolo[macro]), 'IR_corretto_3Y'
                        ].rank(method='first', na_option='bottom', ascending=False) + foglio['ranking_finale'].max()
                    foglio.loc[
                        (foglio['data_di_avvio'] < self.t0_3Y) & (foglio['Best_Worst'] == 'worst')
                        & (foglio['micro_categoria'] != self.classi_metodo_singolo[macro]), 'ranking_finale'
                    ] = foglio.loc[
                        (foglio['data_di_avvio'] < self.t0_3Y) & (foglio['Best_Worst'] == 'worst')
                        & (foglio['micro_categoria'] != self.classi_metodo_singolo[macro]), 'IR_corretto_3Y'
                        ].rank(method='first', na_option='bottom', ascending=False) + foglio['ranking_finale'].max()
                
                # Metodo best-worst doppio
                elif metodo == 'doppio':
                    foglio = Metodi_ranking(foglio).doppio(macro, 'IR', self.t0_1Y, self.t0_3Y, self.intermediario,
                        self.anni_detenzione, self.soluzioni, self.classi_metodo_doppio)
                # Metodo doppio con linearizzazione
                elif metodo == 'doppio_con_linearizzazione':
                    foglio = Metodi_ranking(foglio).doppio_con_linearizzazione(macro, 'IR', self.t0_1Y, self.t0_3Y, self.intermediario,
                        self.anni_detenzione, self.soluzioni, self.classi_metodo_doppio)
                # Metodo classifica
                elif metodo == 'classifica':
                    foglio = Metodi_ranking(foglio).classifica('IR', self.t0_1Y, self.t0_3Y, self.intermediario, self.anni_detenzione)
                # Metodo doppio con linearizzazione
                elif metodo == 'classifica_con_linearizzazione':
                    foglio = Metodi_ranking(foglio).classifica_con_linearizzazione('IR', self.t0_1Y, self.t0_3Y, self.intermediario,
                        self.anni_detenzione)
                # Metodo linearizzazione
                elif metodo == 'linearizzazione':
                    foglio = Metodi_ranking(foglio).linearizzazione('IR', self.t0_1Y, self.t0_3Y, self.intermediario, self.anni_detenzione)
                
                # Cambio formato data
                # foglio['data_di_avvio'] = foglio['data_di_avvio'].dt.strftime('%d/%m/%Y')
                # # Ordinamento finale
                # if self.metodo == 'singolo' or self.metodo == 'doppio':
                #     foglio.sort_values('ranking_finale', ascending=True, inplace=True)
                # elif self.metodo == 'linearizzazione':
                #     foglio.sort_values('ranking_finale', ascending=False, inplace=True)
                #     # Etichetta ND per i fondi senza dati
                #     foglio['ranking_finale_1Y'] = foglio['ranking_finale_1Y'].fillna('ND')
                #     foglio['ranking_finale_3Y'] = foglio['ranking_finale_3Y'].fillna('ND')
                #     foglio['ranking_finale'] = foglio['ranking_finale'].fillna('ND')
                # # Reindex
                # foglio.reset_index(drop=True, inplace=True)

                # Seleziona colonne utili
                # if self.metodo == 'singolo':
                #     foglio = foglio[
                #         ['ISIN', 'valuta', 'nome', 'data_di_avvio', 'Best_Worst', 'micro_categoria', 'ranking_finale',
                #         'Information_Ratio_3Y', 'ranking_IR_3Y', 'quartile_IR_3Y', 'terzile_IR_3Y', 'TEV_3Y', 'commissione',
                #         'IR_corretto_3Y', 'ranking_IR_3Y_corretto', 'quartile_IR_corretto_3Y', 'terzile_IR_corretto_3Y',
                #         'Information_Ratio_1Y', 'ranking_IR_1Y', 'quartile_IR_1Y', 'terzile_IR_1Y', 'TEV_1Y', 'commissione',
                #         'IR_corretto_1Y', 'ranking_IR_1Y_corretto', 'quartile_IR_corretto_1Y', 'terzile_IR_corretto_1Y', 'SFDR', 'note']
                #     ]
                # elif self.metodo == 'doppio':
                #     if self.intermediario == 'CRV':
                #         foglio = foglio[
                #             ['ISIN', 'valuta', 'nome', 'data_di_avvio', 'micro_categoria', 'Best_Worst_3Y', 'grado_gestione_3Y', 
                #             'Best_Worst_1Y', 'grado_gestione_1Y', 'ranking_per_grado_3Y', 'ranking_per_grado_1Y', 'punteggio_finale',
                #             'Information_Ratio_3Y', 'TEV_3Y', 'commissione', 'IR_corretto_3Y', 'Information_Ratio_1Y', 'TEV_1Y',
                #             'commissione', 'IR_corretto_1Y', 'SFDR', 'note']
                #         ]
                #     else:
                #         foglio = foglio[
                #             ['ISIN', 'valuta', 'nome', 'data_di_avvio', 'micro_categoria', 'Best_Worst_3Y', 'grado_gestione_3Y', 
                #             'Best_Worst_1Y', 'grado_gestione_1Y', 'ranking_per_grado_3Y', 'ranking_per_grado_1Y', 'ranking_finale',
                #             'Information_Ratio_3Y', 'TEV_3Y', 'commissione', 'IR_corretto_3Y', 'Information_Ratio_1Y', 'TEV_1Y',
                #             'commissione', 'IR_corretto_1Y', 'SFDR', 'note']
                #         ]
                # elif self.metodo == 'linearizzazione':
                #     foglio = foglio[
                #         ['ISIN', 'valuta', 'nome', 'data_di_avvio', 'micro_categoria', 'podio', 'ranking_finale', 'ranking_finale_3Y',
                #         'ranking_finale_1Y', 'Information_Ratio_3Y', 'TEV_3Y', 'commissione', 'IR_corretto_3Y', 'Information_Ratio_1Y', 
                #         'TEV_1Y', 'commissione', 'IR_corretto_1Y', 'note']
                #     ]
                # Crea foglio
                foglio.to_excel(writer, sheet_name=macro)

            elif macro in self.SOR_DSR:
                # if self.metodo == 'singolo' or self.metodo == 'doppio':
                # Metodo classifica
                if metodo == 'classifica':
                    foglio = Metodi_ranking(foglio).classifica('SO', self.t0_1Y, self.t0_3Y, self.intermediario, self.anni_detenzione)              
                # Metodo doppio con linearizzazione
                elif metodo == 'classifica_con_linearizzazione':
                    foglio = Metodi_ranking(foglio).classifica_con_linearizzazione('SO', self.t0_1Y, self.t0_3Y, self.intermediario,
                        self.anni_detenzione)
                # Metodo linearizzazione
                elif metodo == 'linearizzazione':
                    foglio = Metodi_ranking(foglio).linearizzazione('SO', self.t0_1Y, self.t0_3Y, self.intermediario, self.anni_detenzione)
                # Crea foglio
                foglio.to_excel(writer, sheet_name=macro)
            
            elif macro in self.SHA_VOL:
                # if self.metodo == 'singolo' or self.metodo == 'doppio':
                # Metodo classifica
                if metodo == 'classifica':
                    foglio = Metodi_ranking(foglio).classifica('SH', self.t0_1Y, self.t0_3Y, self.intermediario, self.anni_detenzione)
                # Metodo doppio con linearizzazione
                elif metodo == 'classifica_con_linearizzazione':
                    foglio = Metodi_ranking(foglio).classifica_con_linearizzazione('SH', self.t0_1Y, self.t0_3Y, self.intermediario,
                        self.anni_detenzione)
                # Metodo linearizzazione
                elif metodo == 'linearizzazione':
                    foglio = Metodi_ranking(foglio).linearizzazione('SH', self.t0_1Y, self.t0_3Y, self.intermediario, self.anni_detenzione)
                # Crea foglio
                foglio.to_excel(writer, sheet_name=macro)

            elif macro in self.PER_VOL:
                # TODO: per la liquidità attiva il metodo doppio, per la liquidità straniera un altro
                # Metodo best-worst singolo
                if self.metodo == 'singolo' or (self.metodo == 'doppio' and macro == 'LIQ_FOR'):
                    # Rank PERF_1Y
                    foglio['ranking_PERF_1Y'] = foglio.loc[
                        (foglio['data_di_avvio'] < self.t0_1Y) & (foglio['Perf_1Y'].notnull()), 'Perf_1Y'
                    ].rank(method='first', na_option='bottom', ascending=False)
                    # Quartile PERF_1Y
                    foglio['quartile_PERF_1Y'] = foglio.loc[(foglio['data_di_avvio'] < self.t0_1Y) & (foglio['Perf_1Y'].notnull()), 'Perf_1Y'].apply(lambda x: 'best' if x > foglio['Perf_1Y'].quantile(0.25, interpolation = 'linear') else 'worst')
                    # Terzile PERF_1Y
                    foglio['terzile_PERF_1Y'] = foglio.loc[(foglio['data_di_avvio'] < self.t0_1Y) & (foglio['Perf_1Y'].notnull()), 'Perf_1Y'].apply(lambda x: 'best' if x > foglio['Perf_1Y'].quantile(0.33, interpolation = 'linear') else 'worst')
                    # Creazione PERF_corretto_1Y
                    if self.intermediario == 'RAI': # Raiffeisen specifica gli anni di detenzione per fondo, non in maniera universale.
                        foglio['PERF_corretto_1Y'] = (df['Perf_1Y'] / 100) - (df['commissione'] / df['anni_detenzione'])
                    else:
                        foglio['PERF_corretto_1Y'] = (df['Perf_1Y'] / 100) - (df['commissione'] / self.anni_detenzione)
                    # Rank PERF_corretto_1Y
                    foglio['ranking_PERF_1Y_corretto'] = foglio.loc[(foglio['data_di_avvio'] < self.t0_1Y) & (foglio['PERF_corretto_1Y'].notnull()), 'PERF_corretto_1Y'].rank(method='first', na_option='bottom', ascending=False)
                    # Quartile PERF_1Y corretto
                    foglio['quartile_PERF_corretto_1Y'] = foglio.loc[(foglio['data_di_avvio'] < self.t0_1Y) & (foglio['PERF_corretto_1Y'].notnull()), 'PERF_corretto_1Y'].apply(lambda x: 'best' if x > foglio['PERF_corretto_1Y'].quantile(0.25, interpolation = 'linear') else 'worst')
                    # Terzile PERF_1Y corretto
                    foglio['terzile_PERF_corretto_1Y'] = foglio.loc[(foglio['data_di_avvio'] < self.t0_1Y) & (foglio['PERF_corretto_1Y'].notnull()), 'PERF_corretto_1Y'].apply(lambda x: 'best' if x > foglio['PERF_corretto_1Y'].quantile(0.33, interpolation = 'linear') else 'worst')

                    # Rank PERF_3Y
                    foglio['ranking_PERF_3Y'] = foglio.loc[(foglio['data_di_avvio'] < self.t0_3Y) & (foglio['Perf_3Y'].notnull()), 'Perf_3Y'].rank(method='first', na_option='keep', ascending=False)
                    # Note
                    foglio.loc[(foglio['data_di_avvio'] < self.t0_3Y) & foglio['Perf_3Y'].isnull(), 'note'] = 'Ha 3 anni, ma non possiede dati a tre anni.'
                    # foglio.loc[(foglio['data_di_avvio'] > t0_3Y) & foglio['Perf_3Y'].notnull(), 'note'] = 'Non ha 3 anni, ma possiede dati a tre anni.'
                    # Quartile PERF_3Y TOGLI IL BEST_WORST
                    foglio['quartile_PERF_3Y'] = foglio.loc[(foglio['data_di_avvio'] < self.t0_3Y) & (foglio['Perf_3Y'].notnull()), 'Perf_3Y'].apply(lambda x: 'best' if x > foglio['Perf_3Y'].quantile(0.25, interpolation = 'linear') else 'worst')
                    # Terzile PERF_3Y
                    foglio['terzile_PERF_3Y'] = foglio.loc[(foglio['data_di_avvio'] < self.t0_3Y) & (foglio['Perf_3Y'].notnull()), 'Perf_3Y'].apply(lambda x: 'best' if x > foglio['Perf_3Y'].quantile(0.33, interpolation = 'linear') else 'worst')
                    # Creazione PERF_corretto_3Y (la volatilità è già in percentuale)
                    if self.intermediario == 'RAI': # Raiffeisen specifica gli anni di detenzione per fondo, non in maniera universale.
                        foglio['PERF_corretto_3Y'] = (df['Perf_3Y'] / 100) - (df['commissione'] / df['anni_detenzione'])
                    else:
                        foglio['PERF_corretto_3Y'] = (df['Perf_3Y'] / 100) - (df['commissione'] / self.anni_detenzione)
                    # Rank PERF_corretto_3Y
                    foglio['ranking_PERF_3Y_corretto'] = foglio.loc[(foglio['data_di_avvio'] < self.t0_3Y) & (foglio['PERF_corretto_3Y'].notnull()), 'PERF_corretto_3Y'].rank(method='first', na_option='bottom', ascending=False)
                    # Quartile PERF_3Y corretto
                    foglio['quartile_PERF_corretto_3Y'] = foglio.loc[(foglio['data_di_avvio'] < self.t0_3Y) & (foglio['PERF_corretto_3Y'].notnull()), 'PERF_corretto_3Y'].apply(lambda x: 'best' if x > foglio['PERF_corretto_3Y'].quantile(0.25, interpolation = 'linear') else 'worst')
                    # Terzile PERF_3Y corretto
                    foglio['terzile_PERF_corretto_3Y'] = foglio.loc[(foglio['data_di_avvio'] < self.t0_3Y) & (foglio['PERF_corretto_3Y'].notnull()), 'PERF_corretto_3Y'].apply(lambda x: 'best' if x > foglio['PERF_corretto_3Y'].quantile(0.33, interpolation = 'linear') else 'worst')
                
                # Metodo best-worst doppio
                elif metodo == 'doppio':
                    foglio = Metodi_ranking(foglio).doppio(macro, 'PERF', self.t0_1Y, self.t0_3Y, self.intermediario,
                        self.anni_detenzione, self.soluzioni, self.classi_metodo_doppio)
                # Metodo doppio con linearizzazione
                elif metodo == 'doppio_con_linearizzazione':
                    foglio = Metodi_ranking(foglio).doppio_con_linearizzazione(macro, 'PERF', self.t0_1Y, self.t0_3Y, self.intermediario,
                        self.anni_detenzione, self.soluzioni, self.classi_metodo_doppio)
                elif metodo == 'classifica':
                    foglio = Metodi_ranking(foglio).classifica('PERF', self.t0_1Y, self.t0_3Y, self.intermediario, self.anni_detenzione)
                # Metodo linearizzazione
                elif metodo == 'linearizzazione':
                    foglio = Metodi_ranking(foglio).linearizzazione('PERF', self.t0_1Y, self.t0_3Y, self.intermediario, self.anni_detenzione)

                # Cambio formato data
                # foglio['data_di_avvio'] = foglio['data_di_avvio'].dt.strftime('%d/%m/%Y')
                # # Ordinamento finale
                # if self.metodo == 'singolo' or (self.metodo == 'doppio' and macro == 'LIQ_FOR'):
                #     foglio.sort_values('ranking_PERF_3Y_corretto', ascending=True, inplace=True)
                # elif self.metodo == 'doppio':
                #     foglio.sort_values('ranking_finale', ascending=True, inplace=True)
                # elif self.metodo == 'linearizzazione':
                #     foglio.sort_values('ranking_finale', ascending=False, inplace=True)
                #     # Etichetta ND per i fondi senza dati
                #     foglio['ranking_finale_1Y'] = foglio['ranking_finale_1Y'].fillna('ND')
                #     foglio['ranking_finale_3Y'] = foglio['ranking_finale_3Y'].fillna('ND')
                #     foglio['ranking_finale'] = foglio['ranking_finale'].fillna('ND')
                # # Reindex
                # foglio.reset_index(drop=True, inplace=True)



                # Seleziona colonne utili
                # if self.metodo == 'singolo':
                #     foglio = foglio[
                #         ['ISIN', 'valuta', 'nome', 'data_di_avvio', 'micro_categoria', 'Perf_3Y', 'ranking_PERF_3Y',
                #         'quartile_PERF_3Y', 'terzile_PERF_3Y', 'Vol_3Y', 'commissione', 'PERF_corretto_3Y',
                #         'ranking_PERF_3Y_corretto', 'quartile_PERF_corretto_3Y', 'terzile_PERF_corretto_3Y', 'Perf_1Y',
                #         'ranking_PERF_1Y', 'quartile_PERF_1Y', 'terzile_PERF_1Y', 'Vol_1Y', 'commissione', 'PERF_corretto_1Y',
                #         'ranking_PERF_1Y_corretto', 'quartile_PERF_corretto_1Y', 'terzile_PERF_corretto_1Y', 'SFDR', 'note']
                #     ]
                # elif self.metodo == 'doppio' and macro == 'LIQ_FOR':
                #     foglio = foglio[
                #         ['ISIN', 'valuta', 'nome', 'data_di_avvio', 'micro_categoria', 'Perf_3Y', 'Vol_3Y', 'commissione',
                #         'PERF_corretto_3Y', 'ranking_PERF_3Y_corretto', 'Perf_1Y', 'Vol_1Y', 'commissione', 'PERF_corretto_1Y',
                #         'ranking_PERF_1Y_corretto', 'note']
                #     ]
                # elif self.metodo == 'doppio' and macro != 'LIQ_FOR':
                #     ### TEST ###
                #     if self.intermediario == 'CRV':
                #         foglio = foglio[
                #             ['ISIN', 'valuta', 'nome', 'data_di_avvio', 'micro_categoria', 'Best_Worst_3Y', 'grado_gestione_3Y',
                #             'Best_Worst_1Y', 'grado_gestione_1Y', 'ranking_per_grado_3Y', 'ranking_per_grado_1Y', 'punteggio_finale',
                #             'Perf_3Y', 'Vol_3Y', 'commissione', 'PERF_corretto_3Y', 'Perf_1Y', 'Vol_1Y', 'commissione',
                #             'PERF_corretto_1Y', 'note']
                #         ]
                #     #######
                #     else:
                #         foglio = foglio[
                #             ['ISIN', 'valuta', 'nome', 'data_di_avvio', 'micro_categoria', 'Best_Worst_3Y', 'grado_gestione_3Y',
                #             'Best_Worst_1Y', 'grado_gestione_1Y', 'ranking_per_grado_3Y', 'ranking_per_grado_1Y', 'ranking_finale',
                #             'Perf_3Y', 'Vol_3Y', 'commissione', 'PERF_corretto_3Y', 'Perf_1Y', 'Vol_1Y', 'commissione',
                #             'PERF_corretto_1Y', 'note']
                #         ]
                # elif self.metodo == 'linearizzazione':
                #     foglio = foglio[
                #         ['ISIN', 'valuta', 'nome', 'data_di_avvio', 'micro_categoria', 'podio', 'ranking_finale',
                #         'ranking_finale_3Y', 'ranking_finale_1Y', 'Perf_3Y', 'Vol_3Y', 'commissione', 'PERF_corretto_3Y',
                #         'Perf_1Y', 'Vol_1Y', 'commissione', 'PERF_corretto_1Y', 'note']
                #     ]

                # Crea foglio
                foglio.to_excel(writer, sheet_name=macro)

        writer.save()

    def aggiunta_colonne(self):
        """Aggiungi eventuali colonne presenti nel file_catalogo alla fine dei fogli del file di ranking
        """
        # Per ripa controlla il nome della colonna
        if self.intermediario == 'BPPB':
            colonne = ['fondo_a_finestra']
        elif self.intermediario == 'CRV':
            colonne = ['nome']
        # elif self.intermediario == 'RIPA':
        #     colonne = ['fondo_equivalente']
        else:
            return None
        # Carica file_catalogo
        df_input = pd.read_excel(self.directory.joinpath(self.file_catalogo), index_col=None)
        # Carica file di ranking senza specificare un foglio per ottenerli tutti
        df_globale = pd.read_excel(self.file_ranking, sheet_name=None) # read all sheets
        with pd.ExcelWriter(self.file_ranking, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
            for sheet in df_globale.keys():
                df = pd.read_excel(self.file_ranking, sheet_name=sheet, index_col=0)
                for colonna in colonne:
                    # Se la colonna da inserire ha lo stesso nome di una colonna già presente modificala
                    if colonna in df.columns:
                        colonna_label = colonna + '_2'
                    else:
                        colonna_label = colonna
                    df_input.rename(columns = {colonna : colonna_label}, inplace=True)
                    df_input_2 = df_input.loc[df_input['isin'].isin(df['ISIN']), ['isin', colonna_label]]
                    # Unisci i due file togliendo la colonna 'isin'
                    df = pd.merge(left=df, right=df_input_2, how='left', left_on='ISIN', right_on='isin').drop('isin', axis=1)
                df.to_excel(writer, sheet_name=sheet)

    def rank_formatted(self):
        """Formatta il file di ranking per mettere in evidenza le classi blend e l'ordinamento definitivo.
        Formatta le intestazioni del file.
        """
        wb = load_workbook(filename='ranking.xlsx') # carica il file
        # Colora le micro blend
        print('sto formattando il file di ranking...\n')

        for sheet in wb.sheetnames:
            metodo = self.metodi[sheet]
            if metodo == 'singolo':
                if sheet in self.classi_metodo_singolo.keys():
                    foglio = wb[sheet] # attiva foglio
                    for cell in foglio['G']:
                        if cell.value == self.classi_metodo_singolo[sheet]: # filtra per micro blend
                            cell.fill = PatternFill(fgColor="f4b084", fill_type='solid') # colora le micro blend
                    for cell in foglio['H']:
                        cell.alignment = Alignment(horizontal='center')
                        cell.fill = PatternFill(fgColor='215967', fill_type='solid') # colora il ranking finale
                    for cell in foglio['J']:
                        cell.alignment = Alignment(horizontal='center')
                        cell.fill = PatternFill(fgColor='92CDDC', fill_type='solid') # colora il ranking IR_3Y
                    for cell in foglio['P']:
                        cell.alignment = Alignment(horizontal='center')
                        cell.fill = PatternFill(fgColor='31869B', fill_type='solid') # colora il ranking IR_3Y corretto
                    for cell in foglio['T']:
                        cell.alignment = Alignment(horizontal='center')
                        cell.fill = PatternFill(fgColor='DAEEF3', fill_type='solid') # colora il ranking IR_1Y
                    for cell in foglio['Z']:
                        cell.alignment = Alignment(horizontal='center')
                        cell.fill = PatternFill(fgColor='B7DEE8', fill_type='solid') # colora il ranking IR_1Y corretto
                    for cell in foglio['O']:
                        cell.number_format = '0.0000'
                    for cell in foglio['Y']:
                        cell.number_format = '0.0000'
                    for cell in foglio['N']:
                        cell.number_format = numbers.FORMAT_PERCENTAGE_00
                    for cell in foglio['X']:
                        cell.number_format = numbers.FORMAT_PERCENTAGE_00
                # elif sheet in micro_blend_classi_non_a_benchmark:
                else:
                    foglio = wb[sheet] # attiva foglio
                    for cell in foglio['H']:
                        cell.alignment = Alignment(horizontal='center')
                        cell.fill = PatternFill(fgColor='92CDDC', fill_type='solid') # colora il ranking PERF_3Y
                    for cell in foglio['N']:
                        cell.alignment = Alignment(horizontal='center')
                        cell.fill = PatternFill(fgColor='31869B', fill_type='solid') # colora il ranking PERF_3Y corretto
                    for cell in foglio['R']:
                        cell.alignment = Alignment(horizontal='center')
                        cell.fill = PatternFill(fgColor='DAEEF3', fill_type='solid') # colora il ranking PERF_1Y
                    for cell in foglio['X']:
                        cell.alignment = Alignment(horizontal='center')
                        cell.fill = PatternFill(fgColor='B7DEE8', fill_type='solid') # colora il ranking PERF_1Y corretto
                    for cell in foglio['M']:
                        cell.number_format = '0.0000'
                    for cell in foglio['W']:
                        cell.number_format = '0.0000'
                    for cell in foglio['L']:
                        cell.number_format = numbers.FORMAT_PERCENTAGE_00
                    for cell in foglio['V']:
                        cell.number_format = numbers.FORMAT_PERCENTAGE_00
            elif metodo == 'doppio':
                foglio = wb[sheet] # attiva foglio
                for cell in foglio['F']:
                    if cell.value == self.classi_metodo_doppio[sheet]: # filtra per micro blend
                        cell.fill = PatternFill(fgColor="f4b084", fill_type='solid') # colora le micro blend
                for cell in foglio['G']:
                    if cell.value == 'best': # filtra per best a 3 anni
                        cell.fill = PatternFill(fgColor="ffd700", fill_type='solid') # colora i best che comandano a 3 anni
                    elif cell.value == 'worst' and cell.offset(row=0, column=2).value != 'best': # filtra per worst a 3 anni che non sono best ad 1 anno
                        cell.fill = PatternFill(fgColor="cda434", fill_type='solid') # colora i worst che comandano a 3 anni
                for cell in foglio['H']:
                    if cell.value == 'semi_attivo':
                        cell.fill = PatternFill(fgColor="f8787b", fill_type='solid') # colora i fondi semi attivi
                    elif cell.value == 'attivo':
                        cell.fill = PatternFill(fgColor="e9403e", fill_type='solid') # colora i fondi attivi
                    elif cell.value == 'molto_attivo':
                        cell.fill = PatternFill(fgColor="880001", fill_type='solid') # colora i fondi molto attivi
                for cell in foglio['I']:
                    if cell.value == 'best' and cell.offset(row=0, column=-2).value != 'best': # filtra per best ad 1 anno che non sono best a 3 anni
                        cell.fill = PatternFill(fgColor="ffd700", fill_type='solid') # colora i best che comandano ad 1 anno
                    elif cell.value == 'worst' and cell.offset(row=0, column=-2).value != 'best' and cell.offset(row=0, column=-2).value != 'worst': # filtra per worst ad 1 anno che non sono nè best nè worst a 3 anni
                        cell.fill = PatternFill(fgColor="cda434", fill_type='solid') # colora i worst che comandano ad 1 anno
                for cell in foglio['J']:
                    if cell.value == 'semi_attivo':
                        cell.fill = PatternFill(fgColor="f8787b", fill_type='solid') # colora i fondi semi attivi
                    elif cell.value == 'attivo':
                        cell.fill = PatternFill(fgColor="e9403e", fill_type='solid') # colora i fondi attivi
                    elif cell.value == 'molto_attivo':
                        cell.fill = PatternFill(fgColor="880001", fill_type='solid') # colora i fondi molto attivi
                for cell in foglio['K']:
                    cell.alignment = Alignment(horizontal='center')
                for cell in foglio['L']:
                    cell.alignment = Alignment(horizontal='center')
                for cell in foglio['M']:
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(fgColor='4F81BD', fill_type='solid') # colora il ranking finale
                for cell in foglio['N']:
                    cell.number_format = numbers.FORMAT_NUMBER_00
                for cell in foglio['O']:
                    cell.number_format = numbers.FORMAT_NUMBER_00
                for cell in foglio['P']:
                    cell.number_format = numbers.FORMAT_PERCENTAGE_00
                for cell in foglio['Q']:
                    cell.number_format = numbers.FORMAT_NUMBER_00
                for cell in foglio['R']:
                    cell.number_format = numbers.FORMAT_NUMBER_00
                for cell in foglio['S']:
                    cell.number_format = numbers.FORMAT_NUMBER_00
                for cell in foglio['T']:
                    cell.number_format = numbers.FORMAT_PERCENTAGE_00
                for cell in foglio['U']:
                    cell.number_format = numbers.FORMAT_NUMBER_00
            elif metodo == 'doppio_con_linearizzazione':
                foglio = wb[sheet] # attiva foglio
                for cell in foglio['F']:
                    if cell.value == self.classi_metodo_doppio[sheet]: # filtra per micro blend
                        cell.fill = PatternFill(fgColor="f4b084", fill_type='solid') # colora le micro blend
                for cell in foglio['G']:
                    if cell.value == 'best': # filtra per best a 3 anni
                        cell.fill = PatternFill(fgColor="ffd700", fill_type='solid') # colora i best che comandano a 3 anni
                    elif cell.value == 'worst' and cell.offset(row=0, column=2).value != 'best': # filtra per worst a 3 anni che non sono best ad 1 anno
                        cell.fill = PatternFill(fgColor="cda434", fill_type='solid') # colora i worst che comandano a 3 anni
                for cell in foglio['H']:
                    if cell.value == 'semi_attivo':
                        cell.fill = PatternFill(fgColor="f8787b", fill_type='solid') # colora i fondi semi attivi
                    elif cell.value == 'attivo':
                        cell.fill = PatternFill(fgColor="e9403e", fill_type='solid') # colora i fondi attivi
                    elif cell.value == 'molto_attivo':
                        cell.fill = PatternFill(fgColor="880001", fill_type='solid') # colora i fondi molto attivi
                for cell in foglio['I']:
                    if cell.value == 'best' and cell.offset(row=0, column=-2).value != 'best': # filtra per best ad 1 anno che non sono best a 3 anni
                        cell.fill = PatternFill(fgColor="ffd700", fill_type='solid') # colora i best che comandano ad 1 anno
                    elif cell.value == 'worst' and cell.offset(row=0, column=-2).value != 'best' and cell.offset(row=0, column=-2).value != 'worst': # filtra per worst ad 1 anno che non sono nè best nè worst a 3 anni
                        cell.fill = PatternFill(fgColor="cda434", fill_type='solid') # colora i worst che comandano ad 1 anno
                for cell in foglio['J']:
                    if cell.value == 'semi_attivo':
                        cell.fill = PatternFill(fgColor="f8787b", fill_type='solid') # colora i fondi semi attivi
                    elif cell.value == 'attivo':
                        cell.fill = PatternFill(fgColor="e9403e", fill_type='solid') # colora i fondi attivi
                    elif cell.value == 'molto_attivo':
                        cell.fill = PatternFill(fgColor="880001", fill_type='solid') # colora i fondi molto attivi
                for cell in foglio['K']:
                    cell.alignment = Alignment(horizontal='center')
                for cell in foglio['L']:
                    cell.alignment = Alignment(horizontal='center')
                for cell in foglio['M']:
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(fgColor='4F81BD', fill_type='solid') # colora il ranking finale
                for cell in foglio['N']:
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(fgColor='4F81BD', fill_type='solid') # colora il ranking finale
                for cell in foglio['O']:
                    cell.number_format = numbers.FORMAT_NUMBER_00
                for cell in foglio['P']:
                    cell.number_format = numbers.FORMAT_NUMBER_00
                for cell in foglio['Q']:
                    cell.number_format = numbers.FORMAT_PERCENTAGE_00
                for cell in foglio['R']:
                    cell.number_format = numbers.FORMAT_NUMBER_00
                for cell in foglio['S']:
                    cell.number_format = numbers.FORMAT_NUMBER_00
                for cell in foglio['T']:
                    cell.number_format = numbers.FORMAT_NUMBER_00
                for cell in foglio['U']:
                    cell.number_format = numbers.FORMAT_PERCENTAGE_00
                for cell in foglio['V']:
                    cell.number_format = numbers.FORMAT_NUMBER_00
            elif metodo == 'classifica':
                foglio = wb[sheet] # attiva foglio
                for cell in foglio['K']:
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(fgColor='4F81BD', fill_type='solid') # colora il ranking dell'indicatore corretto a 3 anni
                for cell in foglio['G']:
                    cell.number_format = numbers.FORMAT_NUMBER_00
                for cell in foglio['H']:
                    cell.number_format = numbers.FORMAT_NUMBER_00
                for cell in foglio['I']:
                    cell.number_format = numbers.FORMAT_PERCENTAGE_00
                for cell in foglio['J']:
                    cell.number_format = numbers.FORMAT_NUMBER_00
                for cell in foglio['L']:
                    cell.number_format = numbers.FORMAT_NUMBER_00
                for cell in foglio['M']:
                    cell.number_format = numbers.FORMAT_NUMBER_00
                for cell in foglio['N']:
                    cell.number_format = numbers.FORMAT_PERCENTAGE_00
                for cell in foglio['O']:
                    cell.number_format = numbers.FORMAT_NUMBER_00
                for cell in foglio['P']:
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(fgColor='4F81BD', fill_type='solid') # colora il ranking dell'indicatore corretto a 3 anni
            elif metodo == 'classifica_con_linearizzazione':
                foglio = wb[sheet] # attiva foglio
                for cell in foglio['K']:
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(fgColor='4F81BD', fill_type='solid') # colora il ranking dell'indicatore corretto a 3 anni
                for cell in foglio['L']:
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(fgColor='4F81BD', fill_type='solid') # colora il ranking dell'indicatore corretto a 3 anni
                for cell in foglio['G']:
                    cell.number_format = numbers.FORMAT_NUMBER_00
                for cell in foglio['H']:
                    cell.number_format = numbers.FORMAT_NUMBER_00
                for cell in foglio['I']:
                    cell.number_format = numbers.FORMAT_PERCENTAGE_00
                for cell in foglio['J']:
                    cell.number_format = numbers.FORMAT_NUMBER_00
                for cell in foglio['M']:
                    cell.number_format = numbers.FORMAT_NUMBER_00
                for cell in foglio['N']:
                    cell.number_format = numbers.FORMAT_NUMBER_00
                for cell in foglio['O']:
                    cell.number_format = numbers.FORMAT_PERCENTAGE_00
                for cell in foglio['P']:
                    cell.number_format = numbers.FORMAT_NUMBER_00
                for cell in foglio['Q']:
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(fgColor='4F81BD', fill_type='solid') # colora il ranking dell'indicatore corretto a 3 anni
            elif metodo == 'linearizzazione':
                foglio = wb[sheet] # attiva foglio
                for cell in foglio['H']: # colora il ranking finale
                    if cell.value == 'ND':
                        cell.fill = PatternFill(fgColor="595959", fill_type='solid')
                        cell.alignment = Alignment(horizontal='right')
                    else:
                        try:
                            if float(cell.value) <= 3.0: # 1-3 bronzo 4-6 argento 7-9 oro
                                cell.fill = PatternFill(fgColor="cd7f32", fill_type='solid')
                                cell.number_format = '0.0000'
                            elif float(cell.value) <= 6.0: # 1-3 bronzo 4-6 argento 7-9 oro
                                cell.fill = PatternFill(fgColor="c0c0c0", fill_type='solid')
                                cell.number_format = '0.0000'
                            elif float(cell.value) <= 9.1: # 1-3 bronzo 4-6 argento 7-9 oro
                                cell.fill = PatternFill(fgColor="cda434", fill_type='solid')
                                cell.number_format = '0.0000'
                        except:
                            pass
                for cell in foglio['I']: # colora il ranking finale 3Y
                    if cell.value == 'ND':
                        cell.fill = PatternFill(fgColor="595959", fill_type='solid')
                        cell.alignment = Alignment(horizontal='right')
                    else:
                        try:
                            if float(cell.value) <= 3.0: # 1-3 bronzo 4-6 argento 7-9 oro
                                cell.fill = PatternFill(fgColor="cd7f32", fill_type='solid')
                                cell.number_format = '0.0000'
                            elif float(cell.value) <= 6.0: # 1-3 bronzo 4-6 argento 7-9 oro
                                cell.fill = PatternFill(fgColor="c0c0c0", fill_type='solid')
                                cell.number_format = '0.0000'
                            elif float(cell.value) <= 9.1: # 1-3 bronzo 4-6 argento 7-9 oro
                                cell.fill = PatternFill(fgColor="cda434", fill_type='solid')
                                cell.number_format = '0.0000'
                        except:
                            pass
                for cell in foglio['J']: # colora il ranking finale 1Y
                    if cell.value == 'ND':
                        cell.fill = PatternFill(fgColor="595959", fill_type='solid')
                        cell.alignment = Alignment(horizontal='right')
                    else:
                        try:
                            if float(cell.value) <= 3.0: # 1-3 bronzo 4-6 argento 7-9 oro
                                cell.fill = PatternFill(fgColor="cd7f32", fill_type='solid')
                                cell.number_format = '0.0000'
                            elif float(cell.value) <= 6.0: # 1-3 bronzo 4-6 argento 7-9 oro
                                cell.fill = PatternFill(fgColor="c0c0c0", fill_type='solid')
                                cell.number_format = '0.0000'
                            elif float(cell.value) <= 9.1: # 1-3 bronzo 4-6 argento 7-9 oro
                                cell.fill = PatternFill(fgColor="cda434", fill_type='solid')
                                cell.number_format = '0.0000'
                        except:
                            pass
                for cell in foglio['N']:
                    cell.number_format = '0.0000'
                for cell in foglio['R']:
                    cell.number_format = '0.0000'
                for cell in foglio['M']:
                    cell.number_format = numbers.FORMAT_PERCENTAGE_00
                for cell in foglio['Q']:
                    cell.number_format = numbers.FORMAT_PERCENTAGE_00

        # Colora e cambia stile alle intestazioni
        for sheet in wb.sheetnames:
            foglio = wb[sheet]
            for column in foglio[1]:
                if column.value != '':
                    column.font = Font(name='Palatino Linotype', bold=True, italic=True) 
                    column.fill = PatternFill(fgColor="bfbfbf", fill_type='solid')

        # Ripa: gli ISIN, la valuta e i nomi dei fondi sottostanti alle polizze vengono sottolineati di verde
        catalogo = pd.read_excel('catalogo_fondi.xlsx')
        if self.intermediario == 'RIPA':
            for sheet in wb.sheetnames:
                foglio = wb[sheet] # attiva foglio
                for cell in foglio['B']:
                    if cell.value == 'ISIN':
                        continue
                    if catalogo.loc[catalogo['isin'].isin([cell.value]), 'sottostante'].to_list()[0] == 'si':
                        cell.fill = PatternFill(fgColor="92D032", fill_type='solid') # colora l'ISIN
                        cell.offset(row=0, column=1).fill = PatternFill(fgColor="92D032", fill_type='solid') # colora la valuta
                        cell.offset(row=0, column=2).fill = PatternFill(fgColor="92D032", fill_type='solid') # colora il nome
        
        # Ordina fogli
        if self.intermediario == 'BPPB':
            ordine = [
                'AZ_EUR', 'AZ_NA', 'AZ_PAC', 'AZ_EM', 'OBB_EUR_BT', 'OBB_EUR_MLT', 'OBB_EUR_CORP', 'OBB_EM', 'OBB_GLOB',
                'OBB_HY', 'FLEX_BVOL', 'FLEX_MAVOL', 'OPP', 'LIQ'
            ]
            # for _ in wb._sheets:
            #     print(str(_)[12:-2])
            wb._sheets.sort(key=lambda i: ordine.index(str(i)[12:-2]))
        elif self.intermediario == 'BPL':
            ordine = [
                'AZ_EUR', 'AZ_NA', 'AZ_PAC', 'AZ_EM', 'AZ_GLOB', 'OBB_EUR_BT', 'OBB_EUR_MLT', 'OBB_EUR_CORP', 'OBB_EUR',
                'OBB_GLOB', 'OBB_USA', 'OBB_EM', 'OBB_HY', 'BIL_MBVOL', 'BIL_AVOL', 'FLEX_PR', 'FLEX_DIN', 'OPP', 'LIQ',
                'LIQ_FOR'
            ]
            wb._sheets.sort(key=lambda i: ordine.index(str(i)[12:-2]))
        elif self.intermediario == 'CRV':
            ordine = [
                'AZ_EUR', 'AZ_NA', 'AZ_PAC', 'AZ_EM', 'AZ_GLOB', 'OBB_EUR_BT', 'OBB_EUR_MLT', 'OBB_EUR_CORP', 'OBB_EM',
                'OBB_GLOB', 'OBB_HY', 'FLEX_PR', 'FLEX_DIN', 'OPP', 'LIQ'
            ]
            wb._sheets.sort(key=lambda i: ordine.index(str(i)[12:-2]))
        elif self.intermediario == 'RIPA':
            ordine = [
                'AZ_EUR', 'AZ_NA', 'AZ_PAC', 'AZ_EM', 'AZ_GLOB', 'AZ_BIO', 'AZ_BDC', 'AZ_FIN', 'AZ_AMB', 'AZ_IMM', 'AZ_IND', 
                'AZ_ECO', 'AZ_SAL', 'AZ_SPU', 'AZ_TEC', 'AZ_TEL', 'AZ_ORO', 'AZ_BEAR', 
                'OBB_EUR_BT', 'OBB_EUR_MLT', 'OBB_EUR', 'OBB_EUR_CORP', 'OBB_GLOB', 'OBB_USA', 'OBB_JAP', 'OBB_EM', 'OBB_HY', 
                'COMM', 'FLEX_PR', 'FLEX_DIN', 'PERF_ASS', 'LIQ', 
            ]
            wb._sheets.sort(key=lambda i: ordine.index(str(i)[12:-2]))
        elif self.intermediario == 'RAI':
            ordine = [
                'AZ_EUR', 'AZ_NA', 'AZ_PAC', 'AZ_EM', 'AZ_GLOB', 'OBB_EUR_BT', 'OBB_EUR_MLT', 'OBB_EUR_CORP', 'OBB_EUR',
                'OBB_USA', 'OBB_GLOB', 'OBB_EM', 'OBB_HY', 'BIL_PR', 'BIL_EQ', 'BIL_AGG', 'FLEX_PR', 'FLEX_DIN', 'OPP',
                'LIQ', 'LIQ_FOR'
            ]
            wb._sheets.sort(key=lambda i: ordine.index(str(i)[12:-2]))

        wb.save(self.file_ranking)

    def aggiunta_prodotti_non_presenti(self):
        """Aggiunta foglio con i prodotti non presenti sulla piattaforma
        """
        df = pd.read_csv(self.directory.joinpath('docs', 'prodotti_non_presenti.csv'), sep=';', index_col=None)
        if self.intermediario == 'CRV':
            df['ranking_finale'] = 'ND'
        with pd.ExcelWriter(self.file_ranking,  engine='openpyxl', mode='a') as writer:
            df.to_excel(writer, sheet_name='NON_IN_PIATTAFORMA')

    def autofit(self):
        """Imposta la miglior lunghezza per le colonne selezionate.
        """
        columns = range(1, 30)
        xls_file = win32com.client.Dispatch("Excel.Application")
        xls_file.visible = False
        wb = xls_file.Workbooks.Open(Filename=self.directory.joinpath('ranking.xlsx').__str__())
        # openpyxl_wb = load_workbook(filename='ranking.xlsx') # carica il file
        for ws in wb.Sheets:
            for _, value in enumerate(columns):
                if value > 0: # la colonna 0 e le negative non esistono
                    ws.Columns(value).AutoFit()
                else:
                    continue
            wb.Save()
        xls_file.DisplayAlerts = False
        wb.Close(SaveChanges=True, Filename=self.file_ranking)
        xls_file.Quit()

    def zip_file(self):
        """
        Crea un file zip contenente il file_ranking e le note.
        """
        print(f'\nsto creando il file zip da inviare a {self.intermediario}...')
        rankZip = zipfile.ZipFile(self.file_zip, 'w')
        rankZip.write(self.file_ranking, compress_type=zipfile.ZIP_DEFLATED)
        rankZip.close()


if __name__ == '__main__':
    start = time.perf_counter()
    _ = Ranking(intermediario='RIPA')
    # _.attività()
    # _.indicatore_BS()
    # _.calcolo_best_worst()
    # _.ranking_per_grado()
    _.merge_completo_liste()
    _.rank()
    _.aggiunta_colonne()
    _.rank_formatted()
    # _.aggiunta_prodotti_non_presenti()
    _.autofit()
    # _.zip_file()
    end = time.perf_counter()
    print("Elapsed time: ", round(end - start, 2), 'seconds')