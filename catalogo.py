import glob
import os
import time
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment


class Catalogo():

    def __str__(self):
        return "Importa il catalogo dei fondi e lo formatta nel file catalogo_fondi.xlsx"

    def __init__(self):
        """
        Arguments:
            directory {str} = percorso in cui trovare il file di input
            directory_catalogo {str} = percorso in cui si trova il catalogo
            directory_input_liste_complete {str} = percorso in cui salvare le liste complete
            file_catalogo {str} = file formattato
        """
        directory = Path().cwd()
        self.directory = directory
        self.directory_catalogo = directory.joinpath('docs', 'input')
        self.directory_input_liste_complete = directory.joinpath('docs', 'import_liste_complete_into_Q')
        self.file_catalogo = "catalogo_fondi.xlsx"

    def rimuovi_testoacapo_e_spazi_da_intestazione(self):
        """
        Rimuove l'allinemanto testo a capo, gli spazi e il segno percentuale nelle colonne di un file excel
        e salva la modifica in un nuovo file.
        """
        filename = os.listdir(self.directory_catalogo)[0]
        print(f"sto togliendo l'opzione 'Testo a capo', gli spazi e il segno percentuale nelle colonne del file '{filename}'...")
        wb = load_workbook(filename=self.directory_catalogo.joinpath(filename))
        ws = wb.worksheets[0]
        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
            for _ in range(0, ws.max_column):
                if ws[row[_].coordinate].value is not None:
                    ws[row[_].coordinate].value = str(ws[row[_].coordinate].value).replace(' ','_').replace('\n','').replace('%','')
                ws[row[_].coordinate].alignment = Alignment(wrap_text=False)
        print(f'...salvataggio del file col nome {self.file_catalogo}...')
        if self.file_catalogo != None:
            if self.file_catalogo in os.listdir(self.directory):
                print("...file già presente nella directory!")
                time.sleep(2)
            else:
                wb.save(self.file_catalogo)
                print("salvato!")
    
    def rinomina_colonne(self):
        """Rinomina le colonne del file catalogo con nomi standard."""
        etichette = {'isin' : 'isin', 'nome' : 'nome', 'descrizione' : 'nome', 'name' : 'nome', 'valuta' : 'valuta', 'divisa' : 'valuta', 
            'currency' : 'valuta', 'commissione' : 'commissione', 'commissioni' : 'commissione', 'oneri' : 'commissione', 
            'finestra' : 'fondo_a_finestra', 'arco_temporale_in_anni' : 'anni_detenzione'}
        df = pd.read_excel(self.file_catalogo)
        print(f"\nColonne originali del file 'catalogo_fondi': {df.columns.values}")
        df.columns = [column.lower() for column in df.columns.values]
        for col in df.columns:
            for key, value in etichette.items():
                if col.__contains__(key):
                    index = df.columns.get_loc(col)
                    df.rename(columns={df.columns[index]:value}, inplace=True)
        print(f"Colonne aggiornate: {df.columns.values}")
        df.to_excel(self.file_catalogo, index=False)

    def rimuovi_spazi(self, *colonne):
        """
        Rimuovi gli spazi nelle osservazioni delle colonne di un file excel

        Parameters:
        colonne(tuple) : tuple di colonne del file in cui togliere gli spazi
        """

        df = pd.read_excel(self.file_catalogo)
        for colonna in colonne:
            print(f"\n...sto togliendo gli spazi nella colonna {colonna}")
            df[colonna] = df[colonna].str.strip()
            df[colonna] = df[colonna].str.replace(" ", "")
            print('tolti!\n')
        df.to_excel(self.file_catalogo, index=False)

    def rimuovi_duplicati(self, *colonne):
        """
        Rimuovi i valori duplicati nelle colonne di un file excel

        Parameters:
        colonne(tuple) : tuple di colonne del file in cui rimuovere i duplicati
        """
        df = pd.read_excel(self.file_catalogo)
        for colonna in colonne:
            print(f"\n...sto rimuovendo i valori multipli nella colonna {colonna}")
            duplicates = df[df.duplicated(subset=[colonna])]
            print(f"I valori duplicati sono:\n {duplicates}") if not duplicates.empty else print('Non ci sono valori duplicati')
            print(f"\nIl catalogo aveva {len(df)} fondi...")
            df.drop_duplicates(subset=[colonna], inplace=True) 
            print(f"...ora ne ha {len(df)}\n")
        df.to_excel(self.file_catalogo, index=False)

    def tronca_valore(self, carattere, *colonne):
        """
        Tronca i valori delle colonne al numero di caratteri desiderato

        Parameters:
        carattere(int): carattere oltre il quale troncare la stringa
        colonne(tuple) : tuple di colonne da trasformare
        """

        df = pd.read_excel(self.file_catalogo)
        for colonna in colonne:
            print(f"\n...la colonna {colonna} ora contiene valori con soli {carattere} caratteri\n")
            df[colonna] = df[colonna].str[:3]
        df.to_excel(self.file_catalogo, index=False)

    def letter_case(self, formato, *colonne):
        """
        Trasforma i valori nelle colonne di un file excel nel formato desiderato

        Parameters:
        formato(str): formato in cui convertire la colonna di un file excel [Valori : upper - lower - title]
        colonne(tuple) : tuple di colonne da trasformare
        """

        df = pd.read_excel(self.file_catalogo)
        for colonna in colonne:
            if formato == 'upper':
                print(f"\n...la colonna {colonna} ora contiene valori in maiuscolo\n")
                df[colonna] = df[colonna].str.upper()
            elif formato == 'lower':
                print(f"\n...la colonna {colonna} ora contiene valori in minuscolo\n")
                df[colonna] = df[colonna].str.lower()
            elif formato == 'camel':
                print(f"\n...la colonna {colonna} ora contiene valori con l'iniziale in maiuscolo\n")
                df[colonna] = df[colonna].str.title()
        df.to_excel(self.file_catalogo, index=False)
    
    def string_percentage_to_float(self, *colonne):
        """
        Trasforma i valori percentuali nelle colonne di un file excel in valori numerici decimali non percentuali,
        sostituendo la virgola con il punto.

        Parameters:
        colonne(tuple) : tuple di colonne da trasformare
        """

        df = pd.read_excel(self.file_catalogo)
        for colonna in colonne:
            print(f"\n...sto trasformando la colonna {colonna} in valore numerico\n")
            df[colonna] = df[colonna].replace('%','', regex=True).replace(',','.', regex=True).astype('float')
        df.to_excel(self.file_catalogo, index=False)
    
    def fix_fee(self, fee_column, commissione_massima):
        """
        Trasforma la colonna che contiene valori troppo grandi, derivani da un'errata formattazione del file di input.
        E.g. commissione : 4% con formato percentuale = 0.04 -> corretto
        E.g. commissione : 4% con formato personalizzato = 4 -> sbagliato, la commissione non può essere del 400%.

        Parameters:
        fee_column(str) : colonna da sistemare.
        commissione_massima(int): commissione massima ipotetica applicata in formato numerico decimale non percentuale.
        """
        df = pd.read_excel(self.file_catalogo)
        print(f"\nI fondi con una commissione sbagliata sono:\n{df[df[fee_column] > commissione_massima]}\n")
        df.loc[df[fee_column] > commissione_massima, fee_column] = df.loc[df[fee_column] > commissione_massima, fee_column] / 100
        df.to_excel(self.file_catalogo, index=False)

    def creazione_liste_complete_input(self):
        """
        Crea file csv, con dimensioni massime pari a 1999 elementi, da importare in Quantalys.it.
        Directory in cui vengono salvati i file : './docs/import_liste_complete_into_Q/'.
        Crea la directory se non esiste.
        """
        if not os.path.exists(self.directory_input_liste_complete):
            os.makedirs(self.directory_input_liste_complete)
        while len(os.listdir(self.directory_input_liste_complete)) != 0:
            print(f"\nCi sono dei file presenti nella cartella di download: {glob.glob(self.directory_input_liste_complete.__str__() + '/*')}\n")
            _ = input('cancella i file prima di proseguire, poi premi enter\n')
        df_cat = pd.read_excel(self.file_catalogo)
        print(f"Lunghezza lista completa: {len(df_cat)} fondi")
        chunks = len(df_cat)//2000 +1 # blocchi da 2000 elementi
        print(f"\nNumero liste totali divise in 1999 elementi: {chunks}")
        for chunk in range(chunks):
            df = df_cat.loc[: , ['isin', 'valuta']]
            df = df.iloc[0 + 1999 * chunk : 1999 + 1999 * chunk]
            df.columns = ['codice isin', 'divisa']
            print(f"Lunghezza lista {chunk}: {len(df)}")
            df.to_csv(self.directory_input_liste_complete.joinpath('lista_completa_' + str(chunk) + '.csv'), sep=";", index=False)
        print(f"\nFile presenti nella cartella: {glob.glob(self.directory_input_liste_complete.__str__() + '/*')}\n")


if __name__ == '__main__':
    start = time.perf_counter()
    _ = Catalogo()
    _.rimuovi_testoacapo_e_spazi_da_intestazione()
    _.rinomina_colonne()
    _.rimuovi_spazi('isin', 'valuta')
    _.rimuovi_duplicati('isin')
    _.tronca_valore(3, 'valuta')
    _.letter_case('upper', 'valuta')
    _.string_percentage_to_float('commissione')
    _.fix_fee(fee_column='commissione', commissione_massima=0.0575)
    _.creazione_liste_complete_input()
    end = time.perf_counter()
    print("Elapsed time: ", round(end - start, 2), 'seconds')